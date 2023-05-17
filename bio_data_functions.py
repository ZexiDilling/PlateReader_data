import PySimpleGUI
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
import pandas as pd
import re
from openpyxl_fix_functions import ex_cell
import datetime
from numpy import histogram, arange
from openpyxl.chart import BarChart, Reference
import time

from helper_func import row_col_to_cell

def org(all_data, well):
    """
    Original reading data from the platereader, for bio data

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param well: Witch well is being calculated
    :type well: str
    :return: All the readings from the original readings from the platreader in the same formate as the future methodes
    :rtype: dict
    """
    return all_data["plates"]["original"]["wells"][well]


def norm(all_data, well):
    """
    Normalises the data based on the avg of the minimum for bio data

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param well: Witch well is being calculated
    :type well: str
    :return: Returns a dict over all the values after the have been normalized.
    :rtype: dict
    """
    return all_data["plates"]["original"]["wells"][well] - all_data["calculations"]["original"]["minimum"]["avg"]


def pora(all_data, well):
    """
    Percentage of remaining activity calculation based on the avg of the max of normalised data

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param well: Witch well is being calculated
    :type well: str
    :return: Returns the Percentage of remaining activity based on the normalized data.
    :rtype: dict
    """
    return ((100 * all_data["plates"]["normalised"]["wells"][well]) /
                all_data["calculations"]["normalised"]["max"]["avg"])


def pora_internal(all_data, well):
    """
    Percentage of remaining activity calculation based on the avg of the max of normalised data
    This should properly be deleted as it is not needed...

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param well: Witch well is being calculated
    :type well: str
    :return: Returns the Percentage of remaining activity based on the original data.
    :rtype: dict
    """
    return ((100 * all_data["plates"]["normalised"]["wells"][well]) /
            all_data["calculations"]["original"]["max"]["avg"])


def z_prime_calculator(all_data, method):
    """
    Calculate Z-prime

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param method: The heading for the calculations that is not avg and stdev
    :type method: str
    :return: Returns the Z-Prime value
    :rtype: int
    """
    # OLD set-up
    # return 1 - ((3 * (all_data["calculations"][method]["max"]["stdev"] +
    #             (all_data["calculations"][method]["minimum"]["stdev"]))) /
    #             abs(all_data["calculations"][method]["max"]["avg"] +
    #             (all_data["calculations"][method]["minimum"]["avg"])))

    # Get the max average and standard deviation, and min average and standard deviation from the all_data dictionary
    max_avg = all_data["calculations"][method]["max"]["avg"]
    max_stdev = all_data["calculations"][method]["max"]["stdev"]
    min_avg = all_data["calculations"][method]["minimum"]["avg"]
    min_stdev = all_data["calculations"][method]["minimum"]["stdev"]

    # Calculate the result
    result = 1 - ((3 * (max_stdev + min_stdev)) / abs(max_avg + min_avg))

    # Return the result
    return result


def state_mapping(config, ws, translate_wells_to_cells, plate, init_row, free_col, temp_dict, methode):
    """
    Colour in the state of the wells and write a guide in the site to translate
    Might need to re-writes this modul, to exclude temp_dict, as it should not be needed, as all the information should
    come from the plate layout

    :param config: The config file
    :type config: configparser.ConfigParser
    :param ws: Witch worksheet that mappings is conected too
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param translate_wells_to_cells: A dict containing what well in the plate belongs to what cell in the excel file
    :type translate_wells_to_cells: dict
    :param plate: Plate layout for what well have what state
    :type plate: dict
    :param init_row: Row to start writing to
    :type init_row: int
    :param free_col: What column should be free / the first column after the last column used for the plate data
    :type free_col: int
    :param temp_dict: The data that is being analysed for the state mapping
    :type temp_dict: dict
    :param methode: The Method that is being used
    :type methode: str
    :return: The colouring of the cells, and a reading guide for the colours, in the excel ark
    """
    # Iterate through the wells in the plate
    init_row_start = init_row  # save the initial starting row
    for counter in plate["well_layout"]:
        state = plate["well_layout"][counter]["state"]  # get the state of the current well
        cell_color = config["plate_colouring"][state]  # get the color for the current state
        cell_color = cell_color.replace("#", "")  # remove the "#" symbol from the color
        temp_cell = translate_wells_to_cells[
            plate["well_layout"][counter]["well_id"]]  # get the cell for the current well
        # fill the cell with the color for the current state
        ws[temp_cell].fill = PatternFill("solid", fgColor=cell_color)

    # write the color guide in the worksheet
    for state in temp_dict["plates"][methode]:
        if state != "wells":  # skip the "wells" state
            if init_row_start == init_row:  # write the header only once
                ws[ex_cell(init_row + 1, free_col)] = "well state"
                # ws[ex_cell(init_row + 1, free_col + 1)] = "colour coding"
                # bold the header
                ws[ex_cell(init_row + 1, free_col)].font = Font(b=True)
                # ws[ex_cell(init_row + 1, free_col + 1)].font = Font(b=True)
            # Writes the state and colour each state with the right colour
            temp_colour = config["plate_colouring"][state]
            temp_colour = temp_colour.replace("#", "")
            ws[ex_cell(init_row + 2, free_col)] = state
            ws[ex_cell(init_row + 2, free_col)].fill = PatternFill("solid", fgColor=temp_colour)

            # write the color for the state
            # ws[ex_cell(init_row + 2, free_col + 1)] = config["plate_colouring"][state]
            init_row += 1  # increment the row for the next state


def heatmap(config, ws, pw_dict, translate_wells_to_cells, heatmap_colours):
    """
    Colour code based on values.

    :param ws: The worksheet where the heat map is placed
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param pw_dict: Dict for each well and what state it is (sample, blank, empty....
    :type pw_dict: dict
    :param translate_wells_to_cells: Dict for each well cells value
    :type translate_wells_to_cells: dict
    :return: A heatmap coloured depending on the options, in the excel file.
    """
    # create a list of well ids where the sample is present
    temp_list = [well for well in pw_dict if pw_dict[well] == "sample"]

    # add conditional formatting to the selected wells to create a heatmap
    ws.conditional_formatting.add(
        f"{translate_wells_to_cells[temp_list[0]]}:{translate_wells_to_cells[temp_list[-1]]}",
        ColorScaleRule(
            # set the starting color based on the 10th percentile of the data
            start_type='percentile',
            start_value=10,
            start_color=config["Settings_bio"]["plate_report_heatmap_colours_low"].replace("#", ""),

            # set the middle color based on the 50th percentile of the data
            mid_type='percentile',
            mid_value=50,
            mid_color=config["Settings_bio"]["plate_report_heatmap_colours_mid"].replace("#", ""),

            # set the end color based on the 90th percentile of the data
            end_type='percentile',
            end_value=90,
            end_color=config["Settings_bio"]["plate_report_heatmap_colours_high"].replace("#", "")
        )
    )
    # 2 colours heat mao
    # ws.conditional_formatting.add(f"{translate_wells_to_cells[temp_list[0]]}:"
    #                               f"{translate_wells_to_cells[temp_list[-1]]}",
    #                               ColorScaleRule(start_type="min",
    #                                              start_color=config["colours to hex"][heatmap_colours["start"]],
    #                                              end_type="max",
    #                                              end_color=config["colours to hex"][heatmap_colours["end"]]))


def hit_mapping(ws, temp_dict, pora_threshold, methode, translate_wells_to_cells, free_col, init_row):
    """
    Colour coding a plate depending on hits (the values are selected before hand), and writes the bounderies for the
    hits.

    :param ws: The worksheet where the hit map is
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param temp_dict: A dicts from all_data, that is only containing the values for the plate. These well values are
        the basis for the hits
    :type temp_dict: dict
    :param pora_threshold: The hit map threshold / bounderies
    :type pora_threshold: dict
    :param methode: The method that is being used
    :type methode: str
    :param translate_wells_to_cells: A dict containing what well in the plate belongs to what cell in the excel file
    :type translate_wells_to_cells: dict
    :param free_col: The first free column after the plate.
    :type free_col: int
    :param init_row: The row where the plate can state to be writen in the excel sheet.
    :type init_row: int
    :return: Colouring the cells in the excel file, depending on the boundaries of the hit.
    """

    # Iterate over all wells and fill the cells based on their pora value
    for wells in translate_wells_to_cells:
        if wells not in temp_dict["plates"][methode]["empty"]:
            for split in pora_threshold:
                # Check if the current split is not "colour"
                if split != "colour":
                    # check if it should be added:
                    if pora_threshold[split]["use"]:

                        # Check if the current well's pora value falls within the range of the current split
                        if float(pora_threshold[split]["min"]) < temp_dict["plates"][methode]["wells"][wells] < float(
                                pora_threshold[split]["max"]):
                            # Get the colour for the current split
                            temp_colour = pora_threshold["colour"][split]
                            # Remove the "#" from the colour code
                            cell_color = temp_colour.replace("#", "")
                            temp_cell = translate_wells_to_cells[wells]
                            # Fill the cell with the calculated color
                            ws[temp_cell].fill = PatternFill("solid", fgColor=cell_color)

    # Initialize the row where the colour guide will be written
    write_row = init_row + 1
    # Write the colour guide
    for threshold in pora_threshold:
        if threshold != "colour":
            # Write the threshold
            ws[ex_cell(write_row, free_col)] = threshold
            # Bold and underline the threshold
            ws[ex_cell(write_row, free_col)].font = Font(b=True, underline="single")
            # colour the threshold
            temp_colour = pora_threshold["colour"][threshold]
            temp_colour = temp_colour.replace("#", "")
            # Write the colour for the level
            ws[ex_cell(write_row, free_col)].fill = PatternFill("solid", fgColor=temp_colour)
            for level in pora_threshold[threshold]:
                if level != "use":
                    # Write the level
                    ws[ex_cell(write_row, free_col + 1)] = level

                    # write the value
                    ws[ex_cell(write_row, free_col + 2)] = pora_threshold[threshold][level]
                    write_row += 1


def frequency_writer(ws, headline, data_set, free_col, initial_row, bin_min, bin_max, bin_width, include_outliers):
    """

    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param freq_data:
    :param free_col: What colmn is free in the sheet to start writing on
    :type free_col: int
    :param initial_row: what is the top row to start writing in
    :type initial_row: int
    :param bin_min: Min value for bin. => Min value for frequency data set to look at.
        can either be a value, or be a string, to calculate the bin based on the data set
    type bin_min: int or str
    :param bin_max: Max value for bin. => Max value for frequency data set to look at.
        can either be a value, or be a string, to calculate the bin based on the data set
    type bin_max: int or str
    :param bin_width: How many values is the frequency data split up into.
    :type bin_width: int
    :param include_outliers: if outliers should be included in the data
    :type include_outliers: bool
    :return: frequency data writen up
    """
    col = free_col
    data_bin_min = None
    data_bin_max = None
    bin_min_seperator = 0
    bin_max_seperator = 0


    row = initial_row
    # write headline:
    ws.cell(column=col, row=row, value=headline).font = Font(b=True)
    ws.cell(column=col + 1, row=row, value="Bin Values").font = Font(b=True)
    row += 1


    # Sets bin range to be range of samples values
    if bin_min == "data_set":
        bin_min = min(data_set)
    else:
        data_bin_min = min(data_set)
    if bin_max == "data_set":
        bin_max = max(data_set)
    else:
        data_bin_max = max(data_set)


    # Setup a hist for values below bin_min
    if data_bin_min < bin_min:
        bin_min_width = bin_min - data_bin_min
        min_hist_list_1, min_hist_list_2 = histogram(data_set, bins=arange(data_bin_min, bin_min +
                                                                                bin_width, bin_min_width))
        # Writes in data
        ws.cell(column=col, row=row, value=min_hist_list_1[0])
        ws.cell(column=col + 1, row=row, value=min_hist_list_2[0])

        bin_min_seperator = 1

    temp_list_1, temp_list_2 = \
        histogram(data_set, bins=arange(bin_min, bin_max + bin_width, bin_width))

    # temp_list_1, temp_list_2 = \
    #     histogram(temp_data_set)
    # write data points:
    for data_point in range(len(temp_list_1)):
        ws.cell(column=col, row=row + data_point + bin_min_seperator, value=temp_list_1[data_point])
        ws.cell(column=col + 1, row=row + data_point + bin_min_seperator, value=temp_list_2[data_point])

    # Setup a hist for values above bin_max
    if data_bin_max > bin_max:
        bin_max_width = data_bin_max - bin_max
        max_hist_list_1, max_hist_list_2 = histogram(data_set, bins=arange(bin_max, data_bin_max +
                                                                 bin_width, bin_max_width))

        ws.cell(column=col, row=row + data_point + bin_min_seperator + bin_max_seperator
                , value=max_hist_list_1[0])
        ws.cell(column=col + 1, row=row + data_point + bin_min_seperator + bin_max_seperator
                , value=max_hist_list_2[1])

    col += 1
    max_row = row + data_point + bin_max_seperator + bin_min_seperator

    if include_outliers:
        data_location = {"min_col": free_col, "min_row": initial_row + 1,
                         "max_col": free_col, "max_row": max_row - 1}
        category_location = {"min_col": col, "min_row": initial_row + 1,
                             "max_col": col, "max_row": max_row - 1}
    else:
        data_location = {"min_col": free_col, "min_row": initial_row,
                         "max_col": free_col, "max_row": max_row}
        category_location = {"min_col": col, "min_row": initial_row,
                             "max_col": col, "max_row": max_row}


    return col, data_location, category_location


def bar_chart(ws, title, free_col, initial_row, data_location, category_location):
    """
    Generate a bar chart
    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param title: The headline for the data
    :type title: str
    :param free_col: What colmn is free in the sheet to start writing on
    :type free_col: int
    :param initial_row: what is the top row to start writing in
    :type initial_row: int
    :param data_location: A dict with location (row, col) in the excelsheet, where the data for the
        bar-chart is located
    :type data_location: dict
    :param category_location: A dict with location (row, col) in the excelsheet, where the category for the
        bar-chart is located
    :type category_location: dict
    :param bin_max_seperator: A counter if there are values outside max
    :type bin_max_seperator: int
    :param bin_min_seperator: A counter if there are values outside min
    :type bin_min_seperator : int
    :return:
    """
    col_placement = free_col + 2

    # set up the chart, and the look
    # ToDo move this to settings
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.x_axis.title = "Bins"
    chart.y_axis.title = "Sample amount"
    chart.shape = 4

    data_set = Reference(ws, min_col=data_location["min_col"], min_row=data_location["min_row"],
                         max_col=data_location["max_col"], max_row=data_location["max_row"])
    category_set = Reference(ws, min_col=category_location["min_col"], min_row=category_location["min_row"],
                         max_col=category_location["max_col"], max_row=category_location["max_row"])

    chart.add_data(data_set, titles_from_data=False)
    chart.set_categories(category_set)
    chart_cell = row_col_to_cell(initial_row, col_placement)
    ws.add_chart(chart, chart_cell)

def well_row_col_type(plate_layout):
    """
    Makes two dicts. one for what type/state each well/cell is and a dict that translate each well to a column and row
    values

    :param plate_layout: The plate-layout for what state each well is in (sample, min, max...)
    :type plate_layout: dict
    :return: well_col_row: a dict with a list of values for rows and for columns, well_type: a dict over the
        type/state of each well/cell
    :rtype: dict, dict
    """
    well_type = {}
    well_col_row = {"well_col": [], "well_row": []}
    temp_plate_layout = plate_layout.get("well_layout", plate_layout)

    # Extract the well information
    for counter in temp_plate_layout:
        for keys in temp_plate_layout[counter]:
            # Get well id and column, row information
            if keys == "well_id":
                temp_well_row = re.sub(r"\d+", "", temp_plate_layout[counter][keys])
                temp_well_col = re.sub(r"\D+", "", temp_plate_layout[counter][keys])
                if temp_well_row not in well_col_row["well_row"]:
                    well_col_row["well_row"].append(temp_well_row)
                if temp_well_col not in well_col_row["well_col"]:
                    well_col_row["well_col"].append(temp_well_col)

            # Get well type information
            if keys == "state":
                well_type.setdefault(temp_plate_layout[counter]["state"], []).append(
                    temp_plate_layout[counter]["well_id"])

    return well_col_row, well_type


def original_data_dict(file, plate_layout):
    """
    The original data from the plate reader, loaded in from the excel file

    :param file: the excel file, with the platereaders data
    :type file: str
    :param plate_layout: The platelayout to tell witch cell/well is in witch state (sample, blank, empty....)
    :type plate_layout: dict
    :return:
        - all_data: all_data: A dict over the original data, and what each cell/well is
        - well_col_row: a dict with a list of values for rows and for columns,
        - well_type: a dict over the type/state of each well/cell
        - barcode: The barcode for the plate
    :rtype:
        - dict
        - dict
        - dict
        - str
    """

    # Get well column, row, and type data
    well_col_row, well_type = well_row_col_type(plate_layout)
    try:
        wb = load_workbook(file)
    except InvalidFileException as error:
        print(error)
        return
    sheet = wb.sheetnames[0]
    ws = wb[sheet]

    # Initialize variables for plate type detection
    plate_type_384 = False
    plate_type_1536 = False

    # Initialize dictionary for all data
    all_data = {"plates": {}, "calculations": {}}
    n_rows = len(well_col_row["well_row"])

    # Iterate through each row in the sheet
    for row_index, row in enumerate(ws.values):     #Todo Change this to be a dict, as there can be multiple readings per data. There can be multiple plate readings inside an excel sheet
        for index_row, value in enumerate(row):

            # Get the date
            if value == "Date:":
                date = row[4]

            # Get the barcode
            if value == "Name" and row[1]:
                barcode = row[1]

            # Get the number of rows to skip
            if value == "<>":
                skipped_rows = row_index

            # Check if the plate is 384-well
            if value == "I":
                plate_type_384 = True

            # Check if the plate is 1536-well
            if value == "AA":
                plate_type_1536 = True

    # Load the data into a pandas dataframe, skipping the rows that were specified earlier
    df_plate = pd.read_excel(file, sheet_name=sheet, skiprows=skipped_rows, nrows=n_rows)

    # Convert the dataframe to a dictionary
    df_plate_dict = df_plate.to_dict()

    # Check the size of the plate and display an error message if the size does not match the expected size
    expected_plate_sizes = {
        8: not plate_type_384 or plate_type_1536,
        16: plate_type_384,
        32: plate_type_1536
    }

    if n_rows not in expected_plate_sizes or not expected_plate_sizes[n_rows]:
        PySimpleGUI.PopupError(f"Wrong plate size, data is not {expected_plate_sizes[n_rows]}-wells")
        return None, None, None

    # temp_reading_dict stores the data from the dataframe in a dictionary format
    temp_reading_dict = {}
    for counter, heading in enumerate(df_plate_dict):
        for index, values in enumerate(df_plate_dict[heading]):
            temp_reading_dict[f"{well_col_row['well_row'][index]}{counter}"] = df_plate_dict[heading][values]

    # The code below stores the data in the all_data dictionary
    all_data["plates"]["original"] = {}
    all_data["plates"]["original"]["wells"] = {}
    for state in well_type:
        for well in well_type[state]:
            try:
                all_data["plates"]["original"]["wells"][well] = temp_reading_dict[well]

            # Incase blanked wells are skipped for the reading
            except KeyError:
                all_data["plates"]["original"]["wells"][well] = 1
    try:
        barcode
    except UnboundLocalError:
        barcode = file.split("/")[-1]

    try:
        date
    except UnboundLocalError:
        date = datetime.today()

    return all_data, well_col_row, well_type, barcode, date


def txt_to_xlsx(file, plate_name):

    # get file name from the "file-path".
    file_name = file.split(".")[0]
    # setup the excel sheet:
    wb = Workbook()
    ws = wb.active

    col = 1
    row = 1

    file = open(file, "r")
    lines = file.readlines()

    for row_index, line in enumerate(lines):
        line = line.removesuffix("\n")
        line = line.strip()
        line = " ".join(line.split())
        line = line.split(" ")
        col_index_1 = 0
        for values in line:
            values = values.split("\t")
            for col_index_2, value in enumerate(values):
                ws.cell(column=col + col_index_1 + col_index_2, row=row + row_index, value=value)
            col_index_1 += 1

    file_name = f"{plate_name}.xlsx"
    wb.save(file_name)

    return file_name

if __name__ == "__main__":

    ...
