from openpyxl import Workbook
import numpy as np
from openpyxl.styles import Font

from openpyxl_fix_functions import *
from bio_data_functions import bar_chart, frequency_writer


def _cal_writer_final_report(barcode, ws, all_data, init_row, init_col, report_output):
    """
    Writes the calculations in the combined report for all the plates

    :param barcode: The barcode for the plate
    :type barcode: str
    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param all_data: A dict over all plate date. all the analysed data will be added to this dict
    :type all_data: dict
    :param init_row: The first row to write to in the excel file
    :type init_row: int
    :param init_col: The first row to write to in the excel file
    :type init_col: int
    :param report_output: Gate dict for what information to write. is set in the settings
    :type report_output: dict
    :return: The overview report page in the final report.
    row_counter: The last row writen in
    :rtype: int
    """

    row_counter = init_row

    merge_cells_single_row(barcode, ws, row_counter, init_col, init_col + 2)
    ws.cell(column=init_col, row=row_counter, value=barcode).font = Font(b=True, underline="single")
    row_counter += 1
    for plate_analysed in all_data["calculations"]:
        # Removing other calculations than avg and stdev
        if plate_analysed != "other":
            # Checks to see if the overview of avg and stv should be included
            if report_output[plate_analysed]["overview"]:
                # Writes the analysed method in, if the overview is set to true
                merge_cells_single_row(plate_analysed, ws, row_counter, init_col, init_col + 2, True, "red_line")
                row_counter += 1
                for state in all_data["calculations"][plate_analysed]:
                    if state != "other":
                        if report_output[plate_analysed][state]:

                            ws.cell(column=init_col, row=row_counter, value=state).font = Font(b=True)
                            for calc in all_data["calculations"][plate_analysed][state]:
                                # Writes avg and stdev including values
                                ws.cell(column=init_col + 1, row=row_counter, value=calc)
                                ws.cell(column=init_col + 2, row=row_counter,
                                        value=all_data["calculations"][plate_analysed][state][calc])
                                row_counter += 1
        else:
            if report_output["z_prime"]:
                ws.cell(column=init_col, row=row_counter, value="z-Prime").font = Font(b=True)
                try:
                    ws.cell(column=init_col + 2, row=row_counter,
                            value=all_data["calculations"][plate_analysed]["z_prime"])
                except KeyError:
                    ws.cell(column=init_col + 2, row=row_counter,
                            value="Z-Prime is not calculated for the plates")
                row_counter += 1
            row_counter += 1
    return row_counter


def _well_writer_final_report(ws, hits, final_report_setup, init_row):
    """
    Writes all the wells in a list, in the excel file, for the values that are within the predetermined values.

    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param hits: A dict over all the values that are "hits" and needs to be added
    :type hits: dict
    :param final_report_setup: Gate for what data is added to the report. Is setting in the settings
    :type final_report_setup: dict
    :param init_row: The first row to write to.
    :type init_row: int
    :return: A list of wells in the excel sheet.
    """
    indent_col = 1
    row_counter = init_row

    for barcode in hits:
        # Writes headline for data inserts to see where the data is coming from
        ws.cell(column=indent_col, row=row_counter, value=barcode).font = Font(b=True, underline="single")
        row_counter += 1

        for method in hits[barcode]:
            if final_report_setup["methods"][method]:
                # writes method
                ws.cell(column=indent_col, row=row_counter, value=method).font = Font(b=True)
                row_counter += 1
                for split in hits[barcode][method]:
                    ws.cell(column=indent_col, row=row_counter, value=split).font = Font(b=True)
                    ws.cell(column=indent_col + 1, row=row_counter,
                            value=final_report_setup["pora_threshold"][split]["min"]).font = \
                        Font(underline="single")
                    ws.cell(column=indent_col + 2, row=row_counter,
                            value=final_report_setup["pora_threshold"][split]["max"]).font = \
                        Font(underline="single")
                    row_counter += 1
                    for well in hits[barcode][method][split]:
                        ws.cell(column=indent_col + 1, row=row_counter, value=well)
                        ws.cell(column=indent_col + 2, row=row_counter,
                                value=hits[barcode][method][split][well])
                        row_counter += 1
        indent_col += 4
        row_counter = init_row


def _get_data(all_plate_data, final_report_setup):
    """
    Grabs data that are needed for the different output sheets in the excel file.

    :param all_plate_data: The data for all the plates.
    :type all_plate_data: dict
    :param final_report_setup: The settings for the final report. is set in the settings.
    :type final_report_setup: dict
    :return:
        - temp_hits: All the hits that are within the values
        - data_calc_dict: A dicts over all the calculations and the values
        - plate_counter: The amount of plates that are being analysed
        - all_states: The states that are being used for the analysis
        - all_methods: The methods that are being used for the analysis
    :rtype:
        - dict
        - dict
        - int
        - dict
        - dict
    """
    data_calc_dict = {}
    temp_hits = {}
    plate_counter = 0
    all_states = []
    all_methods = []
    freq_data = {"all_data": []}

    for barcode in all_plate_data:
        plate_counter += 1
        temp_hits[barcode] = {}
        data_calc_dict[barcode] = {}
        for method in all_plate_data[barcode]["plates"]:
            if method != "other":
                if method not in all_methods:
                    all_methods.append(method)
            if final_report_setup["methods"][method]:
                temp_hits[barcode][method] = {}
                freq_data[barcode] = []
                for thresholds in final_report_setup["pora_threshold"]:
                    temp_hits[barcode][method][thresholds] = {}
                print(all_plate_data[barcode]["plates"][method]["wells"])
                for well in all_plate_data[barcode]["plates"][method]["wells"]:
                    if well in all_plate_data[barcode]["plates"][method]["sample"]:
                        temp_well_value = all_plate_data[barcode]["plates"][method]["wells"][well]
                        freq_data[barcode].append(temp_well_value)
                        freq_data["all_data"].append(temp_well_value)
                        for split in final_report_setup["pora_threshold"]:
                            # Check if the specific threshold is include in the report
                            if final_report_setup["pora_threshold"][split]:

                                if float(final_report_setup["pora_threshold"][split]["min"]) < float(temp_well_value) < \
                                        float(final_report_setup["pora_threshold"][split]["max"]):
                                    temp_hits[barcode][method][split][well] = temp_well_value

        for method in all_plate_data[barcode]["calculations"]:
            data_calc_dict[barcode][method] = {}
            if method != "other":
                for state in all_plate_data[barcode]["calculations"][method]:
                    if state not in all_states:
                        all_states.append(state)

                    data_calc_dict[barcode][method][state] = {}
                    for calc in all_plate_data[barcode]["calculations"][method][state]:
                        data_calc_dict[barcode][method][state][calc] = \
                            all_plate_data[barcode]["calculations"][method][state][calc]

            else:
                for other_calc in all_plate_data[barcode]["calculations"][method]:
                    data_calc_dict[barcode][method][other_calc] = \
                        all_plate_data[barcode]["calculations"][method][other_calc]

    return temp_hits, data_calc_dict, plate_counter, all_states, all_methods, freq_data


def _data_writer(ws_matrix, ws_list, data_calc_dict, state, plate_counter, all_methods, use_list, use_max_min):
    """
    Writes all the data, and handles the flow of the data to witch sheet different things are writen in.

    :param ws_matrix: The Worksheet for the data that goes in Matrix formate for calculations
    :type ws_matrix: openpyxl.worksheet.worksheet.Worksheet
    :param ws_list: The worksheet that list calculations and the min/max values for the different once
    :type ws_list: openpyxl.worksheet.worksheet.Worksheet
    :param data_calc_dict: All the data in a dict formate
    :type data_calc_dict: dict
    :param state: What state the data is for (samples, minimum, maximum, blank...)
    :type state: str
    :param plate_counter: The amount of plates that are being analysed
    :type plate_counter: int
    :param all_methods: A list of all the methods
    :type all_methods: list
    :param use_list: If the list data should be added to the report. Is set in the settings.
    :type use_list: bool
    :param use_max_min: If the min_max data should be added to the report. Is set in the settings.
    :type use_max_min: bool
    :return: Values written into the excel sheet for the Matrix, and the list and or min_max depending on settings
    """
    init_row = 4
    init_col = 3
    spacer = 4
    list_spacer_clm = 6

    col_stdev = init_col + plate_counter + spacer
    col_counter = init_col + 1
    row_counter = init_row + 1
    col_stdev_counter = col_stdev + 1
    row_offset = init_row

    list_clm = init_col - 1
    list_row = init_row

    list_row_minmax = init_row

    for index_m, method in enumerate(all_methods):
        temp_avg_dict = {}
        temp_stdev_dict = {}
        mw_col = col_counter
        mw_row = row_counter
        mw_col_stdev = col_stdev_counter

        for barcodes in data_calc_dict:
            # Writes Plate names in row and clm for avg
            ws_matrix.cell(column=init_col - 1, row=row_counter, value=barcodes).font = Font(b=True)
            ws_matrix.cell(column=col_counter, row=row_offset - 1, value=barcodes).font = Font(b=True)

            # Writes Plate names in row and clm for stdev
            ws_matrix.cell(column=col_stdev - 1, row=row_counter, value=barcodes).font = Font(b=True)
            ws_matrix.cell(column=col_stdev_counter, row=row_offset - 1, value=barcodes).font = Font(b=True)

            for index_method, _ in enumerate(data_calc_dict[barcodes]):

                if index_method == 0:
                    # Writes method for avg
                    ws_matrix.cell(column=init_col, row=row_offset - 1, value=method).font = Font(b=True)
                    # Writes method for stdev
                    ws_matrix.cell(column=col_stdev, row=row_offset - 1, value=method).font = Font(b=True)
                    if method != "other":
                        for calc in data_calc_dict[barcodes][method][state]:
                            temp_value = data_calc_dict[barcodes][method][state][calc]
                            # gets avg values
                            if calc == "avg":
                                ws_matrix.cell(column=init_col, row=row_offset, value=calc).font = Font(b=True)
                                ws_matrix.cell(column=init_col, row=row_counter, value=temp_value)
                                ws_matrix.cell(column=col_counter, row=row_offset, value=temp_value)
                                temp_avg_dict[barcodes] = temp_value
                            elif calc == "stdev":
                                ws_matrix.cell(column=col_stdev, row=row_offset, value=calc).font = Font(b=True)
                                ws_matrix.cell(column=col_stdev, row=row_counter, value=temp_value)
                                ws_matrix.cell(column=col_stdev_counter, row=row_offset, value=temp_value)
                                temp_stdev_dict[barcodes] = temp_value
            # Sets offset for next loop, for writing headlines the right place
            col_counter += 1
            row_counter += 1
            col_stdev_counter += 1

        # calculate the % difference between avg for each plate
        _matrix_calculator(ws_matrix, mw_row, mw_col, temp_avg_dict)
        # calculate the % difference between stdev for each plate
        _matrix_calculator(ws_matrix, mw_row, mw_col_stdev, temp_stdev_dict)

        # sortes the dict by size
        temp_avg_dict = _sort_dict(temp_avg_dict)
        temp_stdev_dict = _sort_dict(temp_stdev_dict)

        # writes list of all avg for the different methods
        if use_list:
            _writes_list_of_values(ws_list, list_row, list_clm, temp_avg_dict, "avg", method)
            _writes_list_of_values(ws_list, list_row, list_clm + 3, temp_stdev_dict, "stdev", method)

        # Calculate how much space the list takes
        list_clm = (list_spacer_clm * (len(all_methods))) + 2

        # writes list of all avg for the different methods
        if use_max_min:
            _write_min_max_values(ws_list, list_row_minmax, list_clm, temp_avg_dict, "avg", method)
            _write_min_max_values(ws_list, list_row_minmax, list_clm + 4, temp_stdev_dict, "stdev", method)

        # makes sure that next loop is writen below the first method. One method per row, with avg and stdev for each.
        col_stdev = init_col + plate_counter + spacer
        col_counter = init_col + 1
        row_counter += spacer
        col_stdev_counter = col_stdev + 1
        row_offset += (plate_counter + spacer)
        list_row_minmax += 5
        list_clm = init_col - 1 + (list_spacer_clm * (index_m + 1))


def _matrix_calculator(ws, row, col, temp_data_dict):
    """
    Calculates and writes the values for the Matrix

    :param ws: The Worksheet for the data that goes in Matrix formate for calculations
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param row: The row for the locations of the matrix
    :type row: int
    :param col: The col for the locations of the matrix
    :type col: int
    :param temp_data_dict: The data for the matrix
    :type temp_data_dict: dict
    :return: Written a matrix in the final report for the excel file
    """

    for index_x, value_x in enumerate(temp_data_dict):
        for index_y, value_y in enumerate(temp_data_dict):
            try:
                # Divide the value of `value_x` by the value of `value_y` and multiply by 100
                temp_value = (float(temp_data_dict[value_x]) / float(temp_data_dict[value_y])) * 100
            except (ZeroDivisionError, TypeError):
                # Handle division by zero and TypeError (when a value can't be converted to float)
                temp_value = None
            # Write the calculated value in the cell (column = `col + index_x`, row = `row + index_y`)
            ws.cell(column=col + index_x, row=row + index_y, value=temp_value)


def _sort_dict(temp_data_dict):
    """
    This sorts the dict from lowest values to highest

    :param temp_data_dict: The data for the hit wells.
    :type temp_data_dict: dict
    :return: a sorted dict
    :rtype: dict
    """
    try:
        return {key: value for key, value in sorted(temp_data_dict.items(), key=lambda item: item[1])}
    except TypeError:
        return temp_data_dict


def _writes_list_of_values(ws, row, col, temp_data_dict, item_name, method=None):
    """
    Writes the hit values in a list.
    NEEDS TO ADD COMPOUND DATA TO THIS AT SOME POINT!!!

    :param ws: The Worksheet for the data, for the excel file
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param row: The row where the data is start being written
    :type row: int
    :param col: The col where the data is start being written
    :type col: int
    :param temp_data_dict: The data that needs to be writen. A dict of wells and their values
    :type temp_data_dict: dict
    :param item_name: The name for the data type. avg or stdev
    :type item_name: str
    :param method: What methods the data is from
    :type method: str
    :return: Data written in the excel file.
    """

    # writes method
    if method:
        merge_cells_single_row(method, ws, row - 2, col, col + 1, True, "red_line")

    # writes headline for the list
    ws.cell(column=col, row=row - 1, value="Barcode").font = Font(b=True)
    ws.cell(column=col + 1, row=row - 1, value=item_name).font = Font(b=True)
    # writes list of values and barcode / plate name
    row_counter = 0
    for index, values in enumerate(temp_data_dict):
        ws.cell(column=col, row=row + index, value=values)
        ws.cell(column=col + 1, row=row + index, value=temp_data_dict[values])
        row_counter += 1

    table_name = f"{item_name}_{method}"
    start_row = row - 1
    end_row = start_row + row_counter
    ws.add_table(table_purple(ws, table_name, start_row, col, end_row, col + 1))


def _write_min_max_values(ws, row, col, data_dict, item_name, method=None):
    """
    Writes the min and max values in the list sheet

    :param ws: The Worksheet for the data, for the excel file
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param row: The row where the data is start being written
    :type row: int
    :param col: The col where the data is start being written
    :type col: int
    :param item_name: The name for the data type. avg or stdev
    :type item_name: str
    :param method: What methods the data is from
    :type method: str
    :return: Data written in the excel file.
    """

    # writes method
    if method:
        merge_cells_single_row(method, ws, row - 2, col, col + 2, True, "red_line")

    # writes headlines
    ws.cell(column=col + 1, row=row - 1, value="Barcode").font = Font(b=True)
    ws.cell(column=col + 2, row=row - 1, value=item_name).font = Font(b=True)
    ws.cell(column=col, row=row, value="Maximum").font = Font(b=True)
    ws.cell(column=col, row=row + 1, value="Minimum").font = Font(b=True)

    # removes None values:
    temp_data_dict = {}
    for keys in data_dict:
        if data_dict[keys]:
            temp_data_dict[keys] = data_dict[keys]

    # writes max and barcode / plate name
    temp_dict_max = max(temp_data_dict, key=temp_data_dict.get)

    ws.cell(column=col + 1, row=row, value=temp_dict_max)
    ws.cell(column=col + 2, row=row, value=temp_data_dict[temp_dict_max])

    # writes max and barcode / plate name
    temp_dict_min = min(temp_data_dict, key=temp_data_dict.get)
    ws.cell(column=col + 1, row=row + 1, value=temp_dict_min)
    ws.cell(column=col + 2, row=row + 1, value=temp_data_dict[temp_dict_min])


def _z_prime(ws, data_calc_dict, use_list, use_max_min):
    """
    Writes the Z-Prime report page in the final report

    :param ws: The Worksheet for the data, for the excel file
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param data_calc_dict: A dict over all the Z-Prime values
    :type data_calc_dict: dict
    :param use_list: If the list data should be added to the report. Is set in the settings.
    :type use_list: bool
    :param use_max_min: If the min_max data should be added to the report. Is set in the settings.
    :type use_max_min: bool
    :return: All the data for the Z-Prime in its own sheet in the final report excel file.
    """

    init_row = 2
    init_col = 2
    spacer = 1

    matrix_col = init_col + 8
    matrix_row = init_row + spacer

    col_counter = matrix_col + spacer
    row_counter = matrix_row + spacer

    z_prime_dict = {}

    for barcodes in data_calc_dict:
        # Writes Plate names
        ws.cell(column=matrix_col - 1, row=row_counter, value=barcodes).font = Font(b=True)
        ws.cell(column=col_counter, row=matrix_row - 1, value=barcodes).font = Font(b=True)
        # Writes values for Z-Prime
        try:
            z_prime = data_calc_dict[barcodes]["other_data"]["z_prime"]
        except KeyError:
            z_prime = data_calc_dict[barcodes]["other"]["z_prime"]

        ws.cell(column=matrix_col, row=row_counter, value=z_prime)
        ws.cell(column=col_counter, row=matrix_row, value=z_prime)
        col_counter += 1
        row_counter += 1
        # z_prime_list.append(z_prime)
        z_prime_dict[barcodes] = z_prime

    col_counter = init_col + 1
    row_counter = init_row + 1

    # _matrix_calculator(ws, row_counter, col_counter, z_prime_list)
    _matrix_calculator(ws, matrix_row + 1, matrix_col + 1, z_prime_dict)

    z_prime_dict = _sort_dict(z_prime_dict)
    if use_list:
        _writes_list_of_values(ws, row_counter, init_col, z_prime_dict, "z_prime")
    if use_max_min:
        _write_min_max_values(ws, row_counter, col_counter + 2, z_prime_dict, "z_prime")


def bio_final_report_controller(analyse_method, all_plate_data, output_file, final_report_setup):
    """
    The controller for the flow of data, that writes the final report in an excel file.

    :param analyse_method: What analyse method is being used.
    :type analyse_method: str
    :param all_plate_data: All the data for all the plates, including calculations and well states
    :type all_plate_data: dict
    :param output_file: The name and path for the final report
    :type output_file: str
    :param final_report_setup: The settings for the final report.
    :type final_report_setup: dict
    :return: An excel file ready to be presented... or something..
    """
    wb = Workbook()
    ws_report = wb.active
    ws_report.title = "Full report"
    ws_well_info = wb.create_sheet("Well Info")



    init_row = 2
    init_col = 2
    row = init_row
    col = init_col
    # calc overview:

    for index, barcode in enumerate(all_plate_data):
        row_counter = _cal_writer_final_report(barcode, ws_report, all_plate_data[barcode], row, col,
                                               final_report_setup["calc"])
        # Writes 5 plates horizontal, before changing rows.
        col += 4

        if index % 5 == 0 and index > 0:
            row += row_counter
            col = init_col
    print("calc")

    # gets data:
    temp_hits, data_calc_dict, plate_counter, all_states, all_methods, all_freq_data = \
        _get_data(all_plate_data, final_report_setup)

    print("gets data")

    # write well data
    _well_writer_final_report(ws_well_info, temp_hits, final_report_setup, init_row)

    print("well writen data")
    # writes Matrix of data:

    for states in all_states:
        if states != "other":
            if final_report_setup["data"][states]["matrix"]:

                _data_writer(ws_creator(wb, states, "Matrix"), ws_creator(wb, states, "Lists"), data_calc_dict, states,
                             plate_counter, all_methods, final_report_setup["data"][states]["list"],
                             final_report_setup["data"][states]["max_min"])
    print("Matrix data")
    # writes Z-prime
    if final_report_setup["data"]["z_prime"]["matrix"]:
        ws_z_prime = wb.create_sheet("Z-Prime")
        _z_prime(ws_z_prime, data_calc_dict, final_report_setup["data"]["z_prime"]["list"],
                 final_report_setup["data"]["z_prime"]["max_min"])

    print("z-prime")

    # histograms
    bin_min = 0
    bin_max = 150
    bin_width = 5
    include_outliers = True
    free_col = 1
    initial_row = 1
    ws_histograms = wb.create_sheet("histogram")
    bar_row = initial_row + 34
    bar_col = 1

    for freq_data in all_freq_data:
        headline = freq_data
        data_set = all_freq_data[freq_data]
        free_col, data_location, category_location = \
            frequency_writer(ws_histograms, headline, data_set, free_col, initial_row, bin_min, bin_max, bin_width,
                             include_outliers)

        bar_chart(ws_histograms, headline, bar_col, bar_row, data_location, category_location)
        free_col += 1
        bar_col += 10
        # sets amount of charts per row. 1 chart is aprox 10 cells long
        if bar_col >= 30:
            bar_col = 1
            bar_row += 15

    print(output_file)
    wb.save(output_file)


if __name__ == "__main__":
    final_report_setup = {'methods': {'original': False, 'normalised': False, 'pora': True}, 'analyse': {'sample': True, 'minimum': False, 'max': False, 'empty': False, 'negative': False, 'positive': False, 'blank': False}, 'calc': {'original': {'overview': True, 'sample': False, 'minimum': True, 'max': True, 'empty': False, 'negative': True, 'positive': True, 'blank': False}, 'normalised': {'overview': True, 'sample': False, 'minimum': True, 'max': True, 'empty': False, 'negative': True, 'positive': True, 'blank': False}, 'pora': {'overview': True, 'sample': True, 'minimum': False, 'max': False, 'empty': False, 'negative': False, 'positive': False, 'blank': False}, 'z_prime': True}, 'pora_threshold': {'th_1': {'min': 0.0, 'max': 10.0, 'use': True}, 'th_2': {'min': 10.0, 'max': 20.0, 'use': True}, 'th_3': {'min': 20.0, 'max': 30.0, 'use': True}, 'th_4': {'min': 30.0, 'max': 40.0, 'use': True}, 'th_5': {'min': 40.0, 'max': 50.0, 'use': True}, 'th_6': {'min': 50.0, 'max': 60.0, 'use': True}, 'th_7': {'min': 60.0, 'max': 70.0, 'use': True}, 'th_8': {'min': 70.0, 'max': 80.0, 'use': True}, 'th_9': {'min': 80.0, 'max': 90.0, 'use': True}, 'th_10': {'min': 90.0, 'max': 100.0, 'use': True}}, 'data': {'sample': {'matrix': False, 'list': False, 'max_min': False}, 'minimum': {'matrix': True, 'list': True, 'max_min': True}, 'max': {'matrix': True, 'list': True, 'max_min': True}, 'empty': {'matrix': False, 'list': False, 'max_min': False}, 'negative': {'matrix': True, 'list': True, 'max_min': True}, 'positive': {'matrix': True, 'list': True, 'max_min': True}, 'blank': {'matrix': False, 'list': False, 'max_min': False}, 'z_prime': {'matrix': True, 'list': True, 'max_min': True}}}

    all_plate_data = {'alpha_so1.xlsx': {'plates': {'original': {'wells': {'A1': 0.0003, 'B1': 0.0001, 'C1': 0.0002, 'D1': 0.0001, 'E1': 0.0001, 'F1': 0.0,'G1': 0.0003, 'H1': 0.0001, 'I1': 0.0002, 'J1': 0.0001, 'K1': 0.0002, 'L1': 0.0003, 'M1': 0.0001, 'N1': 0.0, 'O1': 0.0, 'P1': 0.0, 'A2': 0.0001, 'P2': 0.0002, 'A3': 0.0, 'P3': 0.0, 'A4': 0.0001, 'P4': 0.0001, 'A5': 0.0, 'P5': 0.0001, 'A6': 0.0, 'P6': 0.0002, 'A7': 0.0001, 'P7': 0.0001, 'A8': 0.0001, 'P8': 0.0, 'A9': 0.0, 'P9': 0.0001, 'A10': 0.0, 'P10': 0.0, 'A11': 0.0001, 'P11': 0.0001, 'A12': 0.0001, 'P12': 0.0001, 'A13': 0.0001, 'P13': 0.0001, 'A14': 0.0003, 'P14': 0.0001, 'A15': -0.0001, 'P15': -0.0001, 'A16': 0.0, 'P16': 0.0, 'A17': 0.0001, 'P17': 0.0001, 'A18': 0.0001, 'P18': 0.0, 'A19': 0.0, 'P19': 0.0, 'A20': 0.0001, 'P20': 0.0, 'A21': 0.0001, 'P21': 0.0, 'A22': 0.0003, 'P22': 0.0, 'A23': 0.0001, 'B23': 0.0001, 'C23': 0.0, 'D23': 0.0, 'E23': 0.0002, 'F23': 0.0002, 'G23': 0.0002, 'H23': 0.0001, 'I23': 0.0001, 'J23': 0.0002, 'K23': -0.0001, 'L23': 0.0, 'M23': 0.0, 'N23': -0.0002, 'O23': 0.0, 'P23': -0.0001, 'A24': 0.0, 'B24': 0.0002, 'C24': -0.0001, 'D24': 0.0001, 'E24': 0.0001, 'F24': 0.0, 'G24': 0.0, 'H24': 0.0002, 'I24': 0.0001, 'J24': -0.0001, 'K24': 0.0, 'L24': 0.0, 'M24': 0.0, 'N24': 0.0001, 'O24': 0.0, 'P24': 0.0, 'B2': 0.0001, 'C2': 0.0, 'D2': 0.0001, 'E2': 0.0002, 'F2': 0.0002, 'G2': 0.0002, 'H2': 0.0001, 'I2': 0.0002, 'J2': 0.0004, 'K2': 0.0001, 'L2': 0.0001, 'M2': 0.0, 'N2': 0.0001, 'O2': 0.0002, 'B3': 0.0, 'C3': 0.0001, 'D3': 0.0001, 'E3': -0.0001, 'F3': 0.0001, 'G3': 0.0, 'H3': 0.0001, 'I3': 0.0001, 'J3': 0.0, 'K3': 0.0001, 'L3': 0.0001, 'M3': 0.0001, 'N3': 0.0001, 'O3': 0.0001, 'B4': 0.0, 'C4': 0.0001, 'D4': 0.0001, 'E4': 0.0, 'F4': 0.0001, 'G4': 0.0001, 'H4': 0.0002, 'I4': -0.0001, 'J4': 0.0001, 'K4': 0.0001, 'L4': 0.0001, 'M4': 0.0, 'N4': -0.0001, 'O4': 0.0, 'B5': 0.0001, 'C5': 0.0, 'D5': 0.0, 'E5': 0.0001, 'F5': 0.0001, 'G5': 0.0001, 'H5': 0.0, 'I5': 0.0002, 'J5': 0.0001, 'K5': 0.0, 'L5': 0.0001, 'M5': 0.0, 'N5': 0.0001, 'O5': 0.0, 'B6': 0.0, 'C6': 0.0001, 'D6': 0.0, 'E6': 0.0002, 'F6': 0.0, 'G6': 0.0001, 'H6': 0.0, 'I6': 0.0001, 'J6': 0.0001, 'K6': 0.0, 'L6': 0.0001, 'M6': 0.0001, 'N6': 0.0001, 'O6': 0.0, 'B7': 0.0001, 'C7': 0.0, 'D7': -0.0001, 'E7': 0.0, 'F7': 0.0001, 'G7': 0.0001, 'H7': 0.0, 'I7': 0.0001, 'J7': 0.0002, 'K7': 0.0001, 'L7': 0.0001, 'M7': 0.0, 'N7': 0.0, 'O7': 0.0, 'B8': 0.0002, 'C8': 0.0002, 'D8': 0.0001, 'E8': 0.0001, 'F8': 0.0003, 'G8': 0.0001, 'H8': 0.0001, 'I8': 0.0, 'J8': 0.0, 'K8': 0.0, 'L8': 0.0001, 'M8': 0.0, 'N8': 0.0, 'O8': 0.0002, 'B9': 0.0002, 'C9': 0.0001, 'D9': 0.0002, 'E9': 0.0001, 'F9': 0.0, 'G9': 0.0, 'H9': 0.0001, 'I9': 0.0002, 'J9': 0.0002, 'K9': 0.0001, 'L9': 0.0002, 'M9': -0.0001, 'N9': 0.0003, 'O9': 0.0, 'B10': 0.0001, 'C10': 0.0002, 'D10': 0.0, 'E10': 0.0001, 'F10': 0.0001, 'G10': 0.0, 'H10': 0.0002, 'I10': 0.0, 'J10': 0.0001, 'K10': 0.0001, 'L10': -0.0002, 'M10': 0.0, 'N10': 0.0, 'O10': 0.0001, 'B11': 0.0001, 'C11': 0.0001, 'D11': 0.0002, 'E11': 0.0001, 'F11': 0.0003, 'G11': 0.0, 'H11': 0.0003, 'I11': 0.0001, 'J11': 0.0001, 'K11': 0.0001, 'L11': 0.0, 'M11': 0.0, 'N11': 0.0001, 'O11': 0.0, 'B12': 0.0001, 'C12': 0.0001, 'D12': 0.0002, 'E12': 0.0, 'F12': 0.0002, 'G12': 0.0002, 'H12': 0.0, 'I12': 0.0, 'J12': 0.0001, 'K12': 0.0003, 'L12': 0.0001, 'M12': 0.0001, 'N12': 0.0, 'O12': 0.0, 'B13': 0.0001, 'C13': 0.0, 'D13': 0.0, 'E13': 0.0, 'F13': 0.0, 'G13': 0.0, 'H13': 0.0002, 'I13': 0.0, 'J13': 0.0001, 'K13': 0.0001, 'L13': 0.0003, 'M13': 0.0, 'N13': 0.0, 'O13': 0.0002, 'B14': 0.0001, 'C14': 0.0002, 'D14': 0.0002, 'E14': 0.0001, 'F14': 0.0001, 'G14': 0.0002, 'H14': 0.0002, 'I14': 0.0003, 'J14': 0.0003, 'K14': 0.0001, 'L14': 0.0001, 'M14': -0.0001, 'N14': 0.0, 'O14': 0.0001, 'B15': 0.0001, 'C15': 0.0, 'D15': 0.0002, 'E15': 0.0001, 'F15': 0.0001, 'G15': 0.0002, 'H15': 0.0002, 'I15': 0.0001, 'J15': 0.0002, 'K15': 0.0001, 'L15': 0.0, 'M15': 0.0, 'N15': 0.0001, 'O15': 0.0, 'B16': 0.0001, 'C16': 0.0, 'D16': 0.0, 'E16': 0.0001, 'F16': 0.0001, 'G16': 0.0001, 'H16': 0.0, 'I16': 0.0001, 'J16': 0.0, 'K16': 0.0001, 'L16': 0.0, 'M16': -0.0001, 'N16': -0.0001, 'O16': 0.0002, 'B17': 0.0, 'C17': 0.0001, 'D17': 0.0001, 'E17': 0.0001, 'F17': 0.0002, 'G17': 0.0, 'H17': 0.0001, 'I17': 0.0, 'J17': 0.0, 'K17': 0.0, 'L17': 0.0002, 'M17': 0.0, 'N17': -0.0001, 'O17': 0.0, 'B18': 0.0, 'C18': 0.0001, 'D18': 0.0001, 'E18': 0.0001, 'F18': 0.0001, 'G18': 0.0, 'H18': 0.0, 'I18': 0.0002, 'J18': 0.0002, 'K18': 0.0001, 'L18': 0.0003, 'M18': 0.0, 'N18': 0.0001, 'O18': 0.0001, 'B19': 0.0001, 'C19': 0.0002, 'D19': 0.0001, 'E19': 0.0, 'F19': 0.0001, 'G19': 0.0001, 'H19': 0.0002, 'I19': 0.0001, 'J19': 0.0001, 'K19': 0.0001, 'L19': 0.0001, 'M19': 0.0002, 'N19': 0.0, 'O19': 0.0001, 'B20': 0.0001, 'C20': 0.0001, 'D20': 0.0003, 'E20': 0.0001, 'F20': 0.0001, 'G20': 0.0001, 'H20': 0.0002, 'I20': 0.0001, 'J20': 0.0001, 'K20': 0.0001, 'L20': 0.0, 'M20': 0.0, 'N20': 0.0001, 'O20': 0.0001, 'B21': 0.0001, 'C21': 0.0001, 'D21': 0.0, 'E21': 0.0, 'F21': 0.0002, 'G21': 0.0002, 'H21': 0.0, 'I21': 0.0, 'J21': 0.0, 'K21': 0.0001, 'L21': 0.0, 'M21': 0.0, 'N21': 0.0, 'O21': 0.0001, 'B22': 0.0002, 'C22': 0.0001, 'D22': -0.0001, 'E22': 0.0, 'F22': 0.0002, 'G22': 0.0001, 'H22': 0.0001, 'I22': 0.0002, 'J22': -0.0001, 'K22': 0.0002, 'L22': 0.0004, 'M22': 0.0, 'N22': 0.0001, 'O22': 0.0002}, 'empty': ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'A2', 'P2', 'A3', 'P3', 'A4', 'P4', 'A5', 'P5', 'A6', 'P6', 'A7', 'P7', 'A8', 'P8', 'A9', 'P9', 'A10', 'P10', 'A11', 'P11', 'A12', 'P12', 'A13', 'P13', 'A14', 'P14', 'A15', 'P15', 'A16', 'P16', 'A17', 'P17', 'A18', 'P18', 'A19', 'P19', 'A20', 'P20', 'A21', 'P21', 'A22', 'P22', 'A23', 'B23', 'C23', 'D23', 'E23', 'F23', 'G23', 'H23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'A24', 'B24', 'C24', 'D24', 'E24', 'F24', 'G24', 'H24', 'I24', 'J24', 'K24', 'L24', 'M24', 'N24', 'O24', 'P24'], 'minimum': ['B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2', 'L2', 'M2', 'N2', 'O2'], 'max': ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3', 'K3', 'L3', 'M3', 'N3', 'O3'], 'sample': ['B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'O4', 'B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5', 'M5', 'N5', 'O5', 'B6', 'C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'M6', 'N6', 'O6', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7', 'J7', 'K7', 'L7', 'M7', 'N7', 'O7', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8', 'N8', 'O8', 'B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9', 'J9', 'K9', 'L9', 'M9', 'N9', 'O9', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10', 'J10', 'K10', 'L10', 'M10', 'N10', 'O10', 'B11', 'C11', 'D11', 'E11', 'F11', 'G11', 'H11', 'I11', 'J11', 'K11', 'L11', 'M11', 'N11', 'O11', 'B12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12', 'I12', 'J12', 'K12', 'L12', 'M12', 'N12', 'O12', 'B13', 'C13', 'D13', 'E13', 'F13', 'G13', 'H13', 'I13', 'J13', 'K13', 'L13', 'M13', 'N13', 'O13', 'B14', 'C14', 'D14', 'E14', 'F14', 'G14', 'H14', 'I14', 'J14', 'K14', 'L14', 'M14', 'N14', 'O14', 'B15', 'C15', 'D15', 'E15', 'F15', 'G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'B16', 'C16', 'D16', 'E16', 'F16', 'G16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16', 'B17', 'C17', 'D17', 'E17', 'F17', 'G17', 'H17', 'I17', 'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'B18', 'C18', 'D18', 'E18', 'F18', 'G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 'B19', 'C19', 'D19', 'E19', 'F19', 'G19', 'H19', 'I19', 'J19', 'K19', 'L19', 'M19', 'N19', 'O19', 'B20', 'C20', 'D20', 'E20', 'F20', 'G20', 'H20', 'I20', 'J20', 'K20', 'L20', 'M20', 'N20', 'O20', 'B21', 'C21', 'D21', 'E21', 'F21', 'G21', 'H21', 'I21', 'J21', 'K21', 'L21', 'M21', 'N21', 'O21', 'B22', 'C22', 'D22', 'E22', 'F22', 'G22', 'H22', 'I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22']}, 'normalised': {'wells': {'A1': 0.0001571428571428571, 'B1': -4.285714285714286e-05, 'C1': 5.714285714285714e-05, 'D1': -4.285714285714286e-05, 'E1': -4.285714285714286e-05, 'F1': -0.00014285714285714287, 'G1': 0.0001571428571428571, 'H1': -4.285714285714286e-05, 'I1': 5.714285714285714e-05, 'J1': -4.285714285714286e-05, 'K1': 5.714285714285714e-05, 'L1': 0.0001571428571428571, 'M1': -4.285714285714286e-05, 'N1': -0.00014285714285714287, 'O1': -0.00014285714285714287, 'P1': -0.00014285714285714287, 'A2': -4.285714285714286e-05, 'P2': 5.714285714285714e-05, 'A3': -0.00014285714285714287, 'P3': -0.00014285714285714287, 'A4': -4.285714285714286e-05, 'P4': -4.285714285714286e-05, 'A5': -0.00014285714285714287, 'P5': -4.285714285714286e-05, 'A6': -0.00014285714285714287, 'P6': 5.714285714285714e-05, 'A7': -4.285714285714286e-05, 'P7': -4.285714285714286e-05, 'A8': -4.285714285714286e-05, 'P8': -0.00014285714285714287, 'A9': -0.00014285714285714287, 'P9': -4.285714285714286e-05, 'A10': -0.00014285714285714287, 'P10': -0.00014285714285714287, 'A11': -4.285714285714286e-05, 'P11': -4.285714285714286e-05, 'A12': -4.285714285714286e-05, 'P12': -4.285714285714286e-05, 'A13': -4.285714285714286e-05, 'P13': -4.285714285714286e-05, 'A14': 0.0001571428571428571, 'P14': -4.285714285714286e-05, 'A15': -0.00024285714285714286, 'P15': -0.00024285714285714286, 'A16': -0.00014285714285714287, 'P16': -0.00014285714285714287, 'A17': -4.285714285714286e-05, 'P17': -4.285714285714286e-05, 'A18': -4.285714285714286e-05, 'P18': -0.00014285714285714287, 'A19': -0.00014285714285714287, 'P19': -0.00014285714285714287, 'A20': -4.285714285714286e-05, 'P20': -0.00014285714285714287, 'A21': -4.285714285714286e-05, 'P21': -0.00014285714285714287, 'A22': 0.0001571428571428571, 'P22': -0.00014285714285714287, 'A23': -4.285714285714286e-05, 'B23': -4.285714285714286e-05, 'C23': -0.00014285714285714287, 'D23': -0.00014285714285714287, 'E23': 5.714285714285714e-05, 'F23': 5.714285714285714e-05, 'G23': 5.714285714285714e-05, 'H23': -4.285714285714286e-05, 'I23': -4.285714285714286e-05, 'J23': 5.714285714285714e-05, 'K23': -0.00024285714285714286, 'L23': -0.00014285714285714287, 'M23': -0.00014285714285714287, 'N23': -0.0003428571428571429, 'O23': -0.00014285714285714287, 'P23': -0.00024285714285714286, 'A24': -0.00014285714285714287, 'B24': 5.714285714285714e-05, 'C24': -0.00024285714285714286, 'D24': -4.285714285714286e-05, 'E24': -4.285714285714286e-05, 'F24': -0.00014285714285714287, 'G24': -0.00014285714285714287, 'H24': 5.714285714285714e-05, 'I24': -4.285714285714286e-05, 'J24': -0.00024285714285714286, 'K24': -0.00014285714285714287, 'L24': -0.00014285714285714287, 'M24': -0.00014285714285714287, 'N24': -4.285714285714286e-05, 'O24': -0.00014285714285714287, 'P24': -0.00014285714285714287, 'B2': -4.285714285714286e-05, 'C2': -0.00014285714285714287, 'D2': -4.285714285714286e-05, 'E2': 5.714285714285714e-05, 'F2': 5.714285714285714e-05, 'G2': 5.714285714285714e-05, 'H2': -4.285714285714286e-05, 'I2': 5.714285714285714e-05, 'J2': 0.00025714285714285715, 'K2': -4.285714285714286e-05, 'L2': -4.285714285714286e-05, 'M2': -0.00014285714285714287, 'N2': -4.285714285714286e-05, 'O2': 5.714285714285714e-05, 'B3': -0.00014285714285714287, 'C3': -4.285714285714286e-05, 'D3': -4.285714285714286e-05, 'E3': -0.00024285714285714286, 'F3': -4.285714285714286e-05, 'G3': -0.00014285714285714287, 'H3': -4.285714285714286e-05, 'I3': -4.285714285714286e-05, 'J3': -0.00014285714285714287, 'K3': -4.285714285714286e-05, 'L3': -4.285714285714286e-05, 'M3': -4.285714285714286e-05, 'N3': -4.285714285714286e-05, 'O3': -4.285714285714286e-05, 'B4': -0.00014285714285714287, 'C4': -4.285714285714286e-05, 'D4': -4.285714285714286e-05, 'E4': -0.00014285714285714287, 'F4': -4.285714285714286e-05, 'G4': -4.285714285714286e-05, 'H4': 5.714285714285714e-05, 'I4': -0.00024285714285714286, 'J4': -4.285714285714286e-05, 'K4': -4.285714285714286e-05, 'L4': -4.285714285714286e-05, 'M4': -0.00014285714285714287, 'N4': -0.00024285714285714286, 'O4': -0.00014285714285714287, 'B5': -4.285714285714286e-05, 'C5': -0.00014285714285714287, 'D5': -0.00014285714285714287, 'E5': -4.285714285714286e-05, 'F5': -4.285714285714286e-05, 'G5': -4.285714285714286e-05, 'H5': -0.00014285714285714287, 'I5': 5.714285714285714e-05, 'J5': -4.285714285714286e-05, 'K5': -0.00014285714285714287, 'L5': -4.285714285714286e-05, 'M5': -0.00014285714285714287, 'N5': -4.285714285714286e-05, 'O5': -0.00014285714285714287, 'B6': -0.00014285714285714287, 'C6': -4.285714285714286e-05, 'D6': -0.00014285714285714287, 'E6': 5.714285714285714e-05, 'F6': -0.00014285714285714287, 'G6': -4.285714285714286e-05, 'H6': -0.00014285714285714287, 'I6': -4.285714285714286e-05, 'J6': -4.285714285714286e-05, 'K6': -0.00014285714285714287, 'L6': -4.285714285714286e-05, 'M6': -4.285714285714286e-05, 'N6': -4.285714285714286e-05, 'O6': -0.00014285714285714287, 'B7': -4.285714285714286e-05, 'C7': -0.00014285714285714287, 'D7': -0.00024285714285714286, 'E7': -0.00014285714285714287, 'F7': -4.285714285714286e-05, 'G7': -4.285714285714286e-05, 'H7': -0.00014285714285714287, 'I7': -4.285714285714286e-05, 'J7': 5.714285714285714e-05, 'K7': -4.285714285714286e-05, 'L7': -4.285714285714286e-05, 'M7': -0.00014285714285714287, 'N7': -0.00014285714285714287, 'O7': -0.00014285714285714287, 'B8': 5.714285714285714e-05, 'C8': 5.714285714285714e-05, 'D8': -4.285714285714286e-05, 'E8': -4.285714285714286e-05, 'F8': 0.0001571428571428571, 'G8': -4.285714285714286e-05, 'H8': -4.285714285714286e-05, 'I8': -0.00014285714285714287, 'J8': -0.00014285714285714287, 'K8': -0.00014285714285714287, 'L8': -4.285714285714286e-05, 'M8': -0.00014285714285714287, 'N8': -0.00014285714285714287, 'O8': 5.714285714285714e-05, 'B9': 5.714285714285714e-05, 'C9': -4.285714285714286e-05, 'D9': 5.714285714285714e-05, 'E9': -4.285714285714286e-05, 'F9': -0.00014285714285714287, 'G9': -0.00014285714285714287, 'H9': -4.285714285714286e-05, 'I9': 5.714285714285714e-05, 'J9': 5.714285714285714e-05, 'K9': -4.285714285714286e-05, 'L9': 5.714285714285714e-05, 'M9': -0.00024285714285714286, 'N9': 0.0001571428571428571, 'O9': -0.00014285714285714287, 'B10': -4.285714285714286e-05, 'C10': 5.714285714285714e-05, 'D10': -0.00014285714285714287, 'E10': -4.285714285714286e-05, 'F10': -4.285714285714286e-05, 'G10': -0.00014285714285714287, 'H10': 5.714285714285714e-05, 'I10': -0.00014285714285714287, 'J10': -4.285714285714286e-05, 'K10': -4.285714285714286e-05, 'L10': -0.0003428571428571429, 'M10': -0.00014285714285714287, 'N10': -0.00014285714285714287, 'O10': -4.285714285714286e-05, 'B11': -4.285714285714286e-05, 'C11': -4.285714285714286e-05, 'D11': 5.714285714285714e-05, 'E11': -4.285714285714286e-05, 'F11': 0.0001571428571428571, 'G11': -0.00014285714285714287, 'H11': 0.0001571428571428571, 'I11': -4.285714285714286e-05, 'J11': -4.285714285714286e-05, 'K11': -4.285714285714286e-05, 'L11': -0.00014285714285714287, 'M11': -0.00014285714285714287, 'N11': -4.285714285714286e-05, 'O11': -0.00014285714285714287, 'B12': -4.285714285714286e-05, 'C12': -4.285714285714286e-05, 'D12': 5.714285714285714e-05, 'E12': -0.00014285714285714287, 'F12': 5.714285714285714e-05, 'G12': 5.714285714285714e-05, 'H12': -0.00014285714285714287, 'I12': -0.00014285714285714287, 'J12': -4.285714285714286e-05, 'K12': 0.0001571428571428571, 'L12': -4.285714285714286e-05, 'M12': -4.285714285714286e-05, 'N12': -0.00014285714285714287, 'O12': -0.00014285714285714287, 'B13': -4.285714285714286e-05, 'C13': -0.00014285714285714287, 'D13': -0.00014285714285714287, 'E13': -0.00014285714285714287, 'F13': -0.00014285714285714287, 'G13': -0.00014285714285714287, 'H13': 5.714285714285714e-05, 'I13': -0.00014285714285714287, 'J13': -4.285714285714286e-05, 'K13': -4.285714285714286e-05, 'L13': 0.0001571428571428571, 'M13': -0.00014285714285714287, 'N13': -0.00014285714285714287, 'O13': 5.714285714285714e-05, 'B14': -4.285714285714286e-05, 'C14': 5.714285714285714e-05, 'D14': 5.714285714285714e-05, 'E14': -4.285714285714286e-05, 'F14': -4.285714285714286e-05, 'G14': 5.714285714285714e-05, 'H14': 5.714285714285714e-05, 'I14': 0.0001571428571428571, 'J14': 0.0001571428571428571, 'K14': -4.285714285714286e-05, 'L14': -4.285714285714286e-05, 'M14': -0.00024285714285714286, 'N14': -0.00014285714285714287, 'O14': -4.285714285714286e-05, 'B15': -4.285714285714286e-05, 'C15': -0.00014285714285714287, 'D15': 5.714285714285714e-05, 'E15': -4.285714285714286e-05, 'F15': -4.285714285714286e-05, 'G15': 5.714285714285714e-05, 'H15': 5.714285714285714e-05, 'I15': -4.285714285714286e-05, 'J15': 5.714285714285714e-05, 'K15': -4.285714285714286e-05, 'L15': -0.00014285714285714287, 'M15': -0.00014285714285714287, 'N15': -4.285714285714286e-05, 'O15': -0.00014285714285714287, 'B16': -4.285714285714286e-05, 'C16': -0.00014285714285714287, 'D16': -0.00014285714285714287, 'E16': -4.285714285714286e-05, 'F16': -4.285714285714286e-05, 'G16': -4.285714285714286e-05, 'H16': -0.00014285714285714287, 'I16': -4.285714285714286e-05, 'J16': -0.00014285714285714287, 'K16': -4.285714285714286e-05, 'L16': -0.00014285714285714287, 'M16': -0.00024285714285714286, 'N16': -0.00024285714285714286, 'O16': 5.714285714285714e-05, 'B17': -0.00014285714285714287, 'C17': -4.285714285714286e-05, 'D17': -4.285714285714286e-05, 'E17': -4.285714285714286e-05, 'F17': 5.714285714285714e-05, 'G17': -0.00014285714285714287, 'H17': -4.285714285714286e-05, 'I17': -0.00014285714285714287, 'J17': -0.00014285714285714287, 'K17': -0.00014285714285714287, 'L17': 5.714285714285714e-05, 'M17': -0.00014285714285714287, 'N17': -0.00024285714285714286, 'O17': -0.00014285714285714287, 'B18': -0.00014285714285714287, 'C18': -4.285714285714286e-05, 'D18': -4.285714285714286e-05, 'E18': -4.285714285714286e-05, 'F18': -4.285714285714286e-05, 'G18': -0.00014285714285714287, 'H18': -0.00014285714285714287, 'I18': 5.714285714285714e-05, 'J18': 5.714285714285714e-05, 'K18': -4.285714285714286e-05, 'L18': 0.0001571428571428571, 'M18': -0.00014285714285714287, 'N18': -4.285714285714286e-05, 'O18': -4.285714285714286e-05, 'B19': -4.285714285714286e-05, 'C19': 5.714285714285714e-05, 'D19': -4.285714285714286e-05, 'E19': -0.00014285714285714287, 'F19': -4.285714285714286e-05, 'G19': -4.285714285714286e-05, 'H19': 5.714285714285714e-05, 'I19': -4.285714285714286e-05, 'J19': -4.285714285714286e-05, 'K19': -4.285714285714286e-05, 'L19': -4.285714285714286e-05, 'M19': 5.714285714285714e-05, 'N19': -0.00014285714285714287, 'O19': -4.285714285714286e-05, 'B20': -4.285714285714286e-05, 'C20': -4.285714285714286e-05, 'D20': 0.0001571428571428571, 'E20': -4.285714285714286e-05, 'F20': -4.285714285714286e-05, 'G20': -4.285714285714286e-05, 'H20': 5.714285714285714e-05, 'I20': -4.285714285714286e-05, 'J20': -4.285714285714286e-05, 'K20': -4.285714285714286e-05, 'L20': -0.00014285714285714287, 'M20': -0.00014285714285714287, 'N20': -4.285714285714286e-05, 'O20': -4.285714285714286e-05, 'B21': -4.285714285714286e-05, 'C21': -4.285714285714286e-05, 'D21': -0.00014285714285714287, 'E21': -0.00014285714285714287, 'F21': 5.714285714285714e-05, 'G21': 5.714285714285714e-05, 'H21': -0.00014285714285714287, 'I21': -0.00014285714285714287, 'J21': -0.00014285714285714287, 'K21': -4.285714285714286e-05, 'L21': -0.00014285714285714287, 'M21': -0.00014285714285714287, 'N21': -0.00014285714285714287, 'O21': -4.285714285714286e-05, 'B22': 5.714285714285714e-05, 'C22': -4.285714285714286e-05, 'D22': -0.00024285714285714286, 'E22': -0.00014285714285714287, 'F22': 5.714285714285714e-05, 'G22': -4.285714285714286e-05, 'H22': -4.285714285714286e-05, 'I22': 5.714285714285714e-05, 'J22': -0.00024285714285714286, 'K22': 5.714285714285714e-05, 'L22': 0.00025714285714285715, 'M22': -0.00014285714285714287, 'N22': -4.285714285714286e-05, 'O22': 5.714285714285714e-05}, 'empty': ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'A2', 'P2', 'A3', 'P3', 'A4', 'P4', 'A5', 'P5', 'A6', 'P6', 'A7', 'P7', 'A8', 'P8', 'A9', 'P9', 'A10', 'P10', 'A11', 'P11', 'A12', 'P12', 'A13', 'P13', 'A14', 'P14', 'A15', 'P15', 'A16', 'P16', 'A17', 'P17', 'A18', 'P18', 'A19', 'P19', 'A20', 'P20', 'A21', 'P21', 'A22', 'P22', 'A23', 'B23', 'C23', 'D23', 'E23', 'F23', 'G23', 'H23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'A24', 'B24', 'C24', 'D24', 'E24', 'F24', 'G24', 'H24', 'I24', 'J24', 'K24', 'L24', 'M24', 'N24', 'O24', 'P24'], 'minimum': ['B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2', 'L2', 'M2', 'N2', 'O2'], 'max': ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3', 'K3', 'L3', 'M3', 'N3', 'O3'], 'sample': ['B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'O4', 'B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5', 'M5', 'N5', 'O5', 'B6', 'C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'M6', 'N6', 'O6', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7', 'J7', 'K7', 'L7', 'M7', 'N7', 'O7', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8', 'N8', 'O8', 'B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9', 'J9', 'K9', 'L9', 'M9', 'N9', 'O9', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10', 'J10', 'K10', 'L10', 'M10', 'N10', 'O10', 'B11', 'C11', 'D11', 'E11', 'F11', 'G11', 'H11', 'I11', 'J11', 'K11', 'L11', 'M11', 'N11', 'O11', 'B12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12', 'I12', 'J12', 'K12', 'L12', 'M12', 'N12', 'O12', 'B13', 'C13', 'D13', 'E13', 'F13', 'G13', 'H13', 'I13', 'J13', 'K13', 'L13', 'M13', 'N13', 'O13', 'B14', 'C14', 'D14', 'E14', 'F14', 'G14', 'H14', 'I14', 'J14', 'K14', 'L14', 'M14', 'N14', 'O14', 'B15', 'C15', 'D15', 'E15', 'F15', 'G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'B16', 'C16', 'D16', 'E16', 'F16', 'G16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16', 'B17', 'C17', 'D17', 'E17', 'F17', 'G17', 'H17', 'I17', 'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'B18', 'C18', 'D18', 'E18', 'F18', 'G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 'B19', 'C19', 'D19', 'E19', 'F19', 'G19', 'H19', 'I19', 'J19', 'K19', 'L19', 'M19', 'N19', 'O19', 'B20', 'C20', 'D20', 'E20', 'F20', 'G20', 'H20', 'I20', 'J20', 'K20', 'L20', 'M20', 'N20', 'O20', 'B21', 'C21', 'D21', 'E21', 'F21', 'G21', 'H21', 'I21', 'J21', 'K21', 'L21', 'M21', 'N21', 'O21', 'B22', 'C22', 'D22', 'E22', 'F22', 'G22', 'H22', 'I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22']}, 'pora': {'wells': {'A1': -199.99999999999994, 'B1': 54.54545454545454, 'C1': -72.72727272727272, 'D1': 54.54545454545454, 'E1': 54.54545454545454, 'F1': 181.8181818181818, 'G1': -199.99999999999994, 'H1': 54.54545454545454, 'I1': -72.72727272727272, 'J1': 54.54545454545454, 'K1': -72.72727272727272, 'L1': -199.99999999999994, 'M1': 54.54545454545454, 'N1': 181.8181818181818, 'O1': 181.8181818181818, 'P1': 181.8181818181818, 'A2': 54.54545454545454, 'P2': -72.72727272727272, 'A3': 181.8181818181818, 'P3': 181.8181818181818, 'A4': 54.54545454545454, 'P4': 54.54545454545454, 'A5': 181.8181818181818, 'P5': 54.54545454545454, 'A6': 181.8181818181818, 'P6': -72.72727272727272, 'A7': 54.54545454545454, 'P7': 54.54545454545454, 'A8': 54.54545454545454, 'P8': 181.8181818181818, 'A9': 181.8181818181818, 'P9': 54.54545454545454, 'A10': 181.8181818181818, 'P10': 181.8181818181818, 'A11': 54.54545454545454, 'P11': 54.54545454545454, 'A12': 54.54545454545454, 'P12': 54.54545454545454, 'A13': 54.54545454545454, 'P13': 54.54545454545454, 'A14': -199.99999999999994, 'P14': 54.54545454545454, 'A15': 309.09090909090907, 'P15': 309.09090909090907, 'A16': 181.8181818181818, 'P16': 181.8181818181818, 'A17': 54.54545454545454, 'P17': 54.54545454545454, 'A18': 54.54545454545454, 'P18': 181.8181818181818, 'A19': 181.8181818181818, 'P19': 181.8181818181818, 'A20': 54.54545454545454, 'P20': 181.8181818181818, 'A21': 54.54545454545454, 'P21': 181.8181818181818, 'A22': -199.99999999999994, 'P22': 181.8181818181818, 'A23': 54.54545454545454, 'B23': 54.54545454545454, 'C23': 181.8181818181818, 'D23': 181.8181818181818, 'E23': -72.72727272727272, 'F23': -72.72727272727272, 'G23': -72.72727272727272, 'H23': 54.54545454545454, 'I23': 54.54545454545454, 'J23': -72.72727272727272, 'K23': 309.09090909090907, 'L23': 181.8181818181818, 'M23': 181.8181818181818, 'N23': 436.3636363636363, 'O23': 181.8181818181818, 'P23': 309.09090909090907, 'A24': 181.8181818181818, 'B24': -72.72727272727272, 'C24': 309.09090909090907, 'D24': 54.54545454545454, 'E24': 54.54545454545454, 'F24': 181.8181818181818, 'G24': 181.8181818181818, 'H24': -72.72727272727272, 'I24': 54.54545454545454, 'J24': 309.09090909090907, 'K24': 181.8181818181818, 'L24': 181.8181818181818, 'M24': 181.8181818181818, 'N24': 54.54545454545454, 'O24': 181.8181818181818, 'P24': 181.8181818181818, 'B2': 54.54545454545454, 'C2': 181.8181818181818, 'D2': 54.54545454545454, 'E2': -72.72727272727272, 'F2': -72.72727272727272, 'G2': -72.72727272727272, 'H2': 54.54545454545454, 'I2': -72.72727272727272, 'J2': -327.27272727272725, 'K2': 54.54545454545454, 'L2': 54.54545454545454, 'M2': 181.8181818181818, 'N2': 54.54545454545454, 'O2': -72.72727272727272, 'B3': 181.8181818181818, 'C3': 54.54545454545454, 'D3': 54.54545454545454, 'E3': 309.09090909090907, 'F3': 54.54545454545454, 'G3': 181.8181818181818, 'H3': 54.54545454545454, 'I3': 54.54545454545454, 'J3': 181.8181818181818, 'K3': 54.54545454545454, 'L3': 54.54545454545454, 'M3': 54.54545454545454, 'N3': 54.54545454545454, 'O3': 54.54545454545454, 'B4': 181.8181818181818, 'C4': 54.54545454545454, 'D4': 54.54545454545454, 'E4': 181.8181818181818, 'F4': 54.54545454545454, 'G4': 54.54545454545454, 'H4': -72.72727272727272, 'I4': 309.09090909090907, 'J4': 54.54545454545454, 'K4': 54.54545454545454, 'L4': 54.54545454545454, 'M4': 181.8181818181818, 'N4': 309.09090909090907, 'O4': 181.8181818181818, 'B5': 54.54545454545454, 'C5': 181.8181818181818, 'D5': 181.8181818181818, 'E5': 54.54545454545454, 'F5': 54.54545454545454, 'G5': 54.54545454545454, 'H5': 181.8181818181818, 'I5': -72.72727272727272, 'J5': 54.54545454545454, 'K5': 181.8181818181818, 'L5': 54.54545454545454, 'M5': 181.8181818181818, 'N5': 54.54545454545454, 'O5': 181.8181818181818, 'B6': 181.8181818181818, 'C6': 54.54545454545454, 'D6': 181.8181818181818, 'E6': -72.72727272727272, 'F6': 181.8181818181818, 'G6': 54.54545454545454, 'H6': 181.8181818181818, 'I6': 54.54545454545454, 'J6': 54.54545454545454, 'K6': 181.8181818181818, 'L6': 54.54545454545454, 'M6': 54.54545454545454, 'N6': 54.54545454545454, 'O6': 181.8181818181818, 'B7': 54.54545454545454, 'C7': 181.8181818181818, 'D7': 309.09090909090907, 'E7': 181.8181818181818, 'F7': 54.54545454545454, 'G7': 54.54545454545454, 'H7': 181.8181818181818, 'I7': 54.54545454545454, 'J7': -72.72727272727272, 'K7': 54.54545454545454, 'L7': 54.54545454545454, 'M7': 181.8181818181818, 'N7': 181.8181818181818, 'O7': 181.8181818181818, 'B8': -72.72727272727272, 'C8': -72.72727272727272, 'D8': 54.54545454545454, 'E8': 54.54545454545454, 'F8': -199.99999999999994, 'G8': 54.54545454545454, 'H8': 54.54545454545454, 'I8': 181.8181818181818, 'J8': 181.8181818181818, 'K8': 181.8181818181818, 'L8': 54.54545454545454, 'M8': 181.8181818181818, 'N8': 181.8181818181818, 'O8': -72.72727272727272, 'B9': -72.72727272727272, 'C9': 54.54545454545454, 'D9': -72.72727272727272, 'E9': 54.54545454545454, 'F9': 181.8181818181818, 'G9': 181.8181818181818, 'H9': 54.54545454545454, 'I9': -72.72727272727272, 'J9': -72.72727272727272, 'K9': 54.54545454545454, 'L9': -72.72727272727272, 'M9': 309.09090909090907, 'N9': -199.99999999999994, 'O9': 181.8181818181818, 'B10': 54.54545454545454, 'C10': -72.72727272727272, 'D10': 181.8181818181818, 'E10': 54.54545454545454, 'F10': 54.54545454545454, 'G10': 181.8181818181818, 'H10': -72.72727272727272, 'I10': 181.8181818181818, 'J10': 54.54545454545454, 'K10': 54.54545454545454, 'L10': 436.3636363636363, 'M10': 181.8181818181818, 'N10': 181.8181818181818, 'O10': 54.54545454545454, 'B11': 54.54545454545454, 'C11': 54.54545454545454, 'D11': -72.72727272727272, 'E11': 54.54545454545454, 'F11': -199.99999999999994, 'G11': 181.8181818181818, 'H11': -199.99999999999994, 'I11': 54.54545454545454, 'J11': 54.54545454545454, 'K11': 54.54545454545454, 'L11': 181.8181818181818, 'M11': 181.8181818181818, 'N11': 54.54545454545454, 'O11': 181.8181818181818, 'B12': 54.54545454545454, 'C12': 54.54545454545454, 'D12': -72.72727272727272, 'E12': 181.8181818181818, 'F12': -72.72727272727272, 'G12': -72.72727272727272, 'H12': 181.8181818181818, 'I12': 181.8181818181818, 'J12': 54.54545454545454, 'K12': -199.99999999999994, 'L12': 54.54545454545454, 'M12': 54.54545454545454, 'N12': 181.8181818181818, 'O12': 181.8181818181818, 'B13': 54.54545454545454, 'C13': 181.8181818181818, 'D13': 181.8181818181818, 'E13': 181.8181818181818, 'F13': 181.8181818181818, 'G13': 181.8181818181818, 'H13': -72.72727272727272, 'I13': 181.8181818181818, 'J13': 54.54545454545454, 'K13': 54.54545454545454, 'L13': -199.99999999999994, 'M13': 181.8181818181818, 'N13': 181.8181818181818, 'O13': -72.72727272727272, 'B14': 54.54545454545454, 'C14': -72.72727272727272, 'D14': -72.72727272727272, 'E14': 54.54545454545454, 'F14': 54.54545454545454, 'G14': -72.72727272727272, 'H14': -72.72727272727272, 'I14': -199.99999999999994, 'J14': -199.99999999999994, 'K14': 54.54545454545454, 'L14': 54.54545454545454, 'M14': 309.09090909090907, 'N14': 181.8181818181818, 'O14': 54.54545454545454, 'B15': 54.54545454545454, 'C15': 181.8181818181818, 'D15': -72.72727272727272, 'E15': 54.54545454545454, 'F15': 54.54545454545454, 'G15': -72.72727272727272, 'H15': -72.72727272727272, 'I15': 54.54545454545454, 'J15': -72.72727272727272, 'K15': 54.54545454545454, 'L15': 181.8181818181818, 'M15': 181.8181818181818, 'N15': 54.54545454545454, 'O15': 181.8181818181818, 'B16': 54.54545454545454, 'C16': 181.8181818181818, 'D16': 181.8181818181818, 'E16': 54.54545454545454, 'F16': 54.54545454545454, 'G16': 54.54545454545454, 'H16': 181.8181818181818, 'I16': 54.54545454545454, 'J16': 181.8181818181818, 'K16': 54.54545454545454, 'L16': 181.8181818181818, 'M16': 309.09090909090907, 'N16': 309.09090909090907, 'O16': -72.72727272727272, 'B17': 181.8181818181818, 'C17': 54.54545454545454, 'D17': 54.54545454545454, 'E17': 54.54545454545454, 'F17': -72.72727272727272, 'G17': 181.8181818181818, 'H17': 54.54545454545454, 'I17': 181.8181818181818, 'J17': 181.8181818181818, 'K17': 181.8181818181818, 'L17': -72.72727272727272, 'M17': 181.8181818181818, 'N17': 309.09090909090907, 'O17': 181.8181818181818, 'B18': 181.8181818181818, 'C18': 54.54545454545454, 'D18': 54.54545454545454, 'E18': 54.54545454545454, 'F18': 54.54545454545454, 'G18': 181.8181818181818, 'H18': 181.8181818181818, 'I18': -72.72727272727272, 'J18': -72.72727272727272, 'K18': 54.54545454545454, 'L18': -199.99999999999994, 'M18': 181.8181818181818, 'N18': 54.54545454545454, 'O18': 54.54545454545454, 'B19': 54.54545454545454, 'C19': -72.72727272727272, 'D19': 54.54545454545454, 'E19': 181.8181818181818, 'F19': 54.54545454545454, 'G19': 54.54545454545454, 'H19': -72.72727272727272, 'I19': 54.54545454545454, 'J19': 54.54545454545454, 'K19': 54.54545454545454, 'L19': 54.54545454545454, 'M19': -72.72727272727272, 'N19': 181.8181818181818, 'O19': 54.54545454545454, 'B20': 54.54545454545454, 'C20': 54.54545454545454, 'D20': -199.99999999999994, 'E20': 54.54545454545454, 'F20': 54.54545454545454, 'G20': 54.54545454545454, 'H20': -72.72727272727272, 'I20': 54.54545454545454, 'J20': 54.54545454545454, 'K20': 54.54545454545454, 'L20': 181.8181818181818, 'M20': 181.8181818181818, 'N20': 54.54545454545454, 'O20': 54.54545454545454, 'B21': 54.54545454545454, 'C21': 54.54545454545454, 'D21': 181.8181818181818, 'E21': 181.8181818181818, 'F21': -72.72727272727272, 'G21': -72.72727272727272, 'H21': 181.8181818181818, 'I21': 181.8181818181818, 'J21': 181.8181818181818, 'K21': 54.54545454545454, 'L21': 181.8181818181818, 'M21': 181.8181818181818, 'N21': 181.8181818181818, 'O21': 54.54545454545454, 'B22': -72.72727272727272, 'C22': 54.54545454545454, 'D22': 309.09090909090907, 'E22': 181.8181818181818, 'F22': -72.72727272727272, 'G22': 54.54545454545454, 'H22': 54.54545454545454, 'I22': -72.72727272727272, 'J22': 309.09090909090907, 'K22': -72.72727272727272, 'L22': -327.27272727272725, 'M22': 181.8181818181818, 'N22': 54.54545454545454, 'O22': -72.72727272727272}, 'empty': ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'A2', 'P2', 'A3', 'P3', 'A4', 'P4', 'A5', 'P5', 'A6', 'P6', 'A7', 'P7', 'A8', 'P8', 'A9', 'P9', 'A10', 'P10', 'A11', 'P11', 'A12', 'P12', 'A13', 'P13', 'A14', 'P14', 'A15', 'P15', 'A16', 'P16', 'A17', 'P17', 'A18', 'P18', 'A19', 'P19', 'A20', 'P20', 'A21', 'P21', 'A22', 'P22', 'A23', 'B23', 'C23', 'D23', 'E23', 'F23', 'G23', 'H23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'A24', 'B24', 'C24', 'D24', 'E24', 'F24', 'G24', 'H24', 'I24', 'J24', 'K24', 'L24', 'M24', 'N24', 'O24', 'P24'], 'minimum': ['B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2', 'L2', 'M2', 'N2', 'O2'], 'max': ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3', 'K3', 'L3', 'M3', 'N3', 'O3'], 'sample': ['B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'O4', 'B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5', 'M5', 'N5', 'O5', 'B6', 'C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'M6', 'N6', 'O6', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7', 'J7', 'K7', 'L7', 'M7', 'N7', 'O7', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8', 'N8', 'O8', 'B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9', 'J9', 'K9', 'L9', 'M9', 'N9', 'O9', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10', 'J10', 'K10', 'L10', 'M10', 'N10', 'O10', 'B11', 'C11', 'D11', 'E11', 'F11', 'G11', 'H11', 'I11', 'J11', 'K11', 'L11', 'M11', 'N11', 'O11', 'B12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12', 'I12', 'J12', 'K12', 'L12', 'M12', 'N12', 'O12', 'B13', 'C13', 'D13', 'E13', 'F13', 'G13', 'H13', 'I13', 'J13', 'K13', 'L13', 'M13', 'N13', 'O13', 'B14', 'C14', 'D14', 'E14', 'F14', 'G14', 'H14', 'I14', 'J14', 'K14', 'L14', 'M14', 'N14', 'O14', 'B15', 'C15', 'D15', 'E15', 'F15', 'G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'B16', 'C16', 'D16', 'E16', 'F16', 'G16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16', 'B17', 'C17', 'D17', 'E17', 'F17', 'G17', 'H17', 'I17', 'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'B18', 'C18', 'D18', 'E18', 'F18', 'G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 'B19', 'C19', 'D19', 'E19', 'F19', 'G19', 'H19', 'I19', 'J19', 'K19', 'L19', 'M19', 'N19', 'O19', 'B20', 'C20', 'D20', 'E20', 'F20', 'G20', 'H20', 'I20', 'J20', 'K20', 'L20', 'M20', 'N20', 'O20', 'B21', 'C21', 'D21', 'E21', 'F21', 'G21', 'H21', 'I21', 'J21', 'K21', 'L21', 'M21', 'N21', 'O21', 'B22', 'C22', 'D22', 'E22', 'F22', 'G22', 'H22', 'I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22']}}, 'calculations': {'original': {'empty': {'avg': 7.000000000000001e-05, 'stdev': 9.994380443501147e-05, 'pstdev': 9.938701010583716e-05, 'pvariance': 9.877777777777777e-09, 'variance': 9.98876404494382e-09, 'st_dev_%': 142.7768634785878}, 'minimum': {'avg': 0.00014285714285714287, 'stdev': 0.00010163498575623618, 'pstdev': 9.793792286287207e-05, 'pvariance': 9.591836734693879e-09, 'variance': 1.0329670329670331e-08, 'st_dev_%': 71.14449002936533}, 'max': {'avg': 6.428571428571429e-05, 'stdev': 6.333236937766509e-05, 'pstdev': 6.102859818083951e-05, 'pvariance': 3.7244897959183676e-09, 'variance': 4.010989010989011e-09, 'st_dev_%': 98.51701903192347}, 'sample': {'avg': 8.383458646616541e-05, 'stdev': 9.153798514163728e-05, 'pstdev': 9.136575924092731e-05, 'pvariance': 8.347701961671095e-09, 'variance': 8.379202723790608e-09, 'st_dev_%': 109.18880738867945}, 'other': {'S/B': 0.44999999999999996}}, 'normalised': {'empty': {'avg': -7.285714285714286e-05, 'stdev': 9.994380443501147e-05, 'pstdev': 9.938701010583716e-05, 'pvariance': 9.877777777777777e-09, 'variance': 9.98876404494382e-09, 'st_dev_%': -137.177770793153}, 'minimum': {'avg': -3.8721506160196585e-21, 'stdev': 0.00010163498575623618, 'pstdev': 9.793792286287207e-05, 'pvariance': 9.591836734693879e-09, 'variance': 1.0329670329670331e-08, 'st_dev_%': -2.6247683996525665e+18}, 'max': {'avg': -7.857142857142858e-05, 'stdev': 6.333236937766509e-05, 'pstdev': 6.102859818083951e-05, 'pvariance': 3.724489795918367e-09, 'variance': 4.010989010989011e-09, 'st_dev_%': -80.60483375339192}, 'sample': {'avg': -5.902255639097745e-05, 'stdev': 9.153798514163728e-05, 'pstdev': 9.136575924092731e-05, 'pvariance': 8.347701961671095e-09, 'variance': 8.379202723790608e-09, 'st_dev_%': -155.08983469857014}, 'other': {'S/B': 2.029141848108051e+16}}, 'pora': {'empty': {'avg': 92.72727272727272, 'stdev': 127.20120564456005, 'pstdev': 126.49255831652, 'pvariance': 16000.367309458215, 'variance': 16180.146717429656, 'st_dev_%': 137.177770793153}, 'minimum': {'avg': 1.0150610510858574e-15, 'stdev': 129.35361823520967, 'pstdev': 124.64826546183716, 'pvariance': 15537.190082644625, 'variance': 16732.358550540368, 'st_dev_%': 1.2743432338068156e+19}, 'max': {'avg': 100.0, 'stdev': 80.60483375339193, 'pstdev': 77.67276132106846, 'pvariance': 6033.057851239669, 'variance': 6497.139224411951, 'st_dev_%': 80.60483375339193}, 'sample': {'avg': 75.11961722488039, 'stdev': 116.50289018026562, 'pstdev': 116.28369357936202, 'pvariance': 13521.89739245896, 'variance': 13572.923420355033, 'st_dev_%': 155.08983469857014}, 'other': {'S/B': 9.85162418487296e+16}}, 'other': {'z_prime': -5.2987535596580475}}}, 'alpha_so2.xlsx': {'plates': {'original': {'wells': {'A1': 0.0498, 'B1': 0.1127, 'C1': 0.0577, 'D1': 0.0534, 'E1': 0.0901, 'F1': 0.1252, 'G1': 0.154, 'H1': 0.1381, 'I1': 0.1712, 'J1': 0.151, 'K1': 0.0738, 'L1': 0.0744, 'M1': 0.2816, 'N1': 0.353, 'O1': 0.226, 'P1': 0.4497, 'A2': 0.0508, 'P2': 0.2465, 'A3': 0.054, 'P3': 0.1446, 'A4': 0.0498, 'P4': 0.0752, 'A5': 0.0502, 'P5': 0.0927, 'A6': 0.0501, 'P6': 0.1563, 'A7': 0.0535, 'P7': 0.1627, 'A8': 0.0499, 'P8': 0.1321, 'A9': 0.0502, 'P9': 0.1182, 'A10': 0.0516, 'P10': 0.1644, 'A11': 0.0511, 'P11': 0.1894, 'A12': 0.0519, 'P12': 0.1758, 'A13': 0.0489, 'P13': 0.1177, 'A14': 0.0523, 'P14': 0.1724, 'A15': 0.0487, 'P15': 0.1315, 'A16': 0.0504, 'P16': 0.2043, 'A17': 0.0572, 'P17': 0.84, 'A18': 0.0518, 'P18': 1.0067, 'A19': 0.0504, 'P19': 0.9991, 'A20': 0.0643, 'P20': 1.1749, 'A21': 0.0801, 'P21': 0.3036, 'A22': 0.1078, 'P22': 0.2057, 'A23': 0.0648, 'B23': 0.088, 'C23': 0.0717, 'D23': 0.0674, 'E23': 0.1181, 'F23': 0.1504, 'G23': 0.1918, 'H23': 0.1477, 'I23': 0.1566, 'J23': 0.1822, 'K23': 0.1095, 'L23': 0.0851, 'M23': 0.2143, 'N23': 0.1884, 'O23': 0.9578, 'P23': 1.1186, 'A24': 0.0526, 'B24': 0.0896, 'C24': 0.0775, 'D24': 0.0741, 'E24': 0.192, 'F24': 0.1278, 'G24': 0.1391, 'H24': 0.1912, 'I24': 0.1689, 'J24': 0.2693, 'K24': 0.0746, 'L24': 0.068, 'M24': 0.3065, 'N24': 0.1924, 'O24': 0.7754, 'P24': 0.9208, 'B2': 0.0794, 'C2': 0.0702, 'D2': 0.0666, 'E2': 0.0855, 'F2': 0.0796, 'G2': 0.0814, 'H2': 0.1149, 'I2': 0.1155, 'J2': 0.1544, 'K2': 0.0975, 'L2': 0.0791, 'M2': 0.3218, 'N2': 0.3795, 'O2': 0.1894, 'B3': 0.2822, 'C3': 0.0956, 'D3': 0.0882, 'E3': 0.1992, 'F3': 0.3353, 'G3': 0.1551, 'H3': 0.3291, 'I3': 0.1343, 'J3': 0.3759, 'K3': 0.0899, 'L3': 0.083, 'M3': 0.1288, 'N3': 0.2269, 'O3': 0.0845, 'B4': 0.3943, 'C4': 0.2285, 'D4': 0.2204, 'E4': 0.1802, 'F4': 0.4154, 'G4': 0.1729, 'H4': 0.2859, 'I4': 0.2538, 'J4': 0.3037, 'K4': 0.1025, 'L4': 0.1558, 'M4': 0.1176, 'N4': 0.2647, 'O4': 0.1583, 'B5': 0.3987, 'C5': 1.0197, 'D5': 1.4678, 'E5': 0.1982, 'F5': 0.3911, 'G5': 0.133, 'H5': 0.5047, 'I5': 0.1981, 'J5': 0.2977, 'K5': 0.0685, 'L5': 0.0905, 'M5': 0.4836, 'N5': 0.3981, 'O5': 0.1413, 'B6': 0.1269, 'C6': 1.3125, 'D6': 1.4579, 'E6': 0.2499, 'F6': 0.4185, 'G6': 0.2154, 'H6': 0.4055, 'I6': 0.2761, 'J6': 0.32, 'K6': 0.0976, 'L6': 0.096, 'M6': 0.5289, 'N6': 0.6771, 'O6': 0.2544, 'B7': 0.0602, 'C7': 0.1002, 'D7': 0.168, 'E7': 0.2661, 'F7': 0.4672, 'G7': 1.0304, 'H7': 1.0834, 'I7': 0.2548, 'J7': 0.3675, 'K7': 0.1196, 'L7': 0.3669, 'M7': 0.5376, 'N7': 0.5397, 'O7': 0.1639, 'B8': 0.0582, 'C8': 0.1043, 'D8': 0.1044, 'E8': 0.2887, 'F8': 0.4677, 'G8': 1.2021, 'H8': 1.1799, 'I8': 0.3223, 'J8': 0.4005, 'K8': 0.1051, 'L8': 0.0897, 'M8': 0.702, 'N8': 0.4458, 'O8': 0.1643, 'B9': 0.0549, 'C9': 0.087, 'D9': 0.0914, 'E9': 1.1404, 'F9': 1.1273, 'G9': 0.263, 'H9': 0.3878, 'I9': 0.3254, 'J9': 0.3832, 'K9': 0.6256, 'L9': 0.4431, 'M9': 0.1469, 'N9': 0.3282, 'O9': 0.196, 'B10': 0.0529, 'C10': 0.0983, 'D10': 0.0894, 'E10': 0.9167, 'F10': 1.245, 'G10': 0.3668, 'H10': 0.4537, 'I10': 0.4415, 'J10': 0.4371, 'K10': 1.2108, 'L10': 0.6491, 'M10': 0.3005, 'N10': 0.403, 'O10': 0.1937, 'B11': 0.0485, 'C11': 0.1025, 'D11': 0.0956, 'E11': 0.429, 'F11': 0.3904, 'G11': 0.35, 'H11': 0.3942, 'I11': 0.3749, 'J11': 0.4406, 'K11': 0.1336, 'L11': 0.1237, 'M11': 0.8801, 'N11': 0.6721, 'O11': 0.302, 'B12': 0.0588, 'C12': 0.0976, 'D12': 0.1041, 'E12': 0.3905, 'F12': 0.4392, 'G12': 0.4069, 'H12': 0.5075, 'I12': 0.4623, 'J12': 0.5281, 'K12': 0.1239, 'L12': 0.1266, 'M12': 0.8701, 'N12': 0.7313, 'O12': 0.354, 'B13': 0.0488, 'C13': 0.0824, 'D13': 0.1026, 'E13': 0.3985, 'F13': 0.4407, 'G13': 0.3045, 'H13': 0.4068, 'I13': 0.3867, 'J13': 0.4989, 'K13': 0.844, 'L13': 1.5849, 'M13': 0.3085, 'N13': 0.3122, 'O13': 0.3081, 'B14': 0.0486, 'C14': 0.1796, 'D14': 0.0983, 'E14': 0.461, 'F14': 0.4522, 'G14': 0.4511, 'H14': 0.4701, 'I14': 0.5253, 'J14': 0.555, 'K14': 1.3096, 'L14': 1.9899, 'M14': 0.4138, 'N14': 0.4021, 'O14': 0.3496, 'B15': 0.0533, 'C15': 0.0801, 'D15': 0.096, 'E15': 0.4682, 'F15': 0.4702, 'G15': 0.3977, 'H15': 0.4646, 'I15': 0.4929, 'J15': 0.7211, 'K15': 0.1058, 'L15': 0.1124, 'M15': 1.1455, 'N15': 1.1576, 'O15': 0.3131, 'B16': 0.055, 'C16': 0.0829, 'D16': 0.0861, 'E16': 0.3958, 'F16': 0.4035, 'G16': 0.3874, 'H16': 0.4172, 'I16': 0.4615, 'J16': 0.4814, 'K16': 0.1219, 'L16': 0.2011, 'M16': 1.1491, 'N16': 1.1425, 'O16': 0.6065, 'B17': 0.0755, 'C17': 0.0844, 'D17': 0.094, 'E17': 0.4696, 'F17': 0.4103, 'G17': 0.3442, 'H17': 0.4159, 'I17': 0.4591, 'J17': 0.5066, 'K17': 0.1123, 'L17': 0.2327, 'M17': 0.7243, 'N17': 0.868, 'O17': 0.7746, 'B18': 0.0502, 'C18': 0.1102, 'D18': 0.0964, 'E18': 0.408, 'F18': 0.3526, 'G18': 0.3176, 'H18': 0.4423, 'I18': 0.4843, 'J18': 0.5823, 'K18': 0.1288, 'L18': 0.1198, 'M18': 0.9704, 'N18': 1.3964, 'O18': 1.1104, 'B19': 0.0626, 'C19': 0.1005, 'D19': 0.113, 'E19': 0.3628, 'F19': 0.3416, 'G19': 0.3039, 'H19': 0.4667, 'I19': 0.6088, 'J19': 0.5679, 'K19': 0.1671, 'L19': 0.3168, 'M19': 1.2627, 'N19': 0.9954, 'O19': 0.9318, 'B20': 0.0484, 'C20': 0.0721, 'D20': 0.1127, 'E20': 0.3722, 'F20': 0.3233, 'G20': 0.2834, 'H20': 0.3946, 'I20': 0.5224, 'J20': 0.5232, 'K20': 0.1075, 'L20': 0.1651, 'M20': 1.3538, 'N20': 1.323, 'O20': 0.9114, 'B21': 0.052, 'C21': 0.0703, 'D21': 0.0651, 'E21': 0.3086, 'F21': 0.3565, 'G21': 0.2921, 'H21': 0.2858, 'I21': 0.4424, 'J21': 0.5458, 'K21': 0.127, 'L21': 0.2045, 'M21': 0.5039, 'N21': 0.6089, 'O21': 0.5252, 'B22': 0.0849, 'C22': 0.0604, 'D22': 0.0719, 'E22': 0.2955, 'F22': 0.2956, 'G22': 0.3197, 'H22': 0.4269, 'I22': 0.4946, 'J22': 0.523, 'K22': 0.1195, 'L22': 0.1131, 'M22': 0.4249, 'N22': 0.5606, 'O22': 0.4931}, 'empty': ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'A2', 'P2', 'A3', 'P3', 'A4', 'P4', 'A5', 'P5', 'A6', 'P6', 'A7', 'P7', 'A8', 'P8', 'A9', 'P9', 'A10', 'P10', 'A11', 'P11', 'A12', 'P12', 'A13', 'P13', 'A14', 'P14', 'A15', 'P15', 'A16', 'P16', 'A17', 'P17', 'A18', 'P18', 'A19', 'P19', 'A20', 'P20', 'A21', 'P21', 'A22', 'P22', 'A23', 'B23', 'C23', 'D23', 'E23', 'F23', 'G23', 'H23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'A24', 'B24', 'C24', 'D24', 'E24', 'F24', 'G24', 'H24', 'I24', 'J24', 'K24', 'L24', 'M24', 'N24', 'O24', 'P24'], 'minimum': ['B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2', 'L2', 'M2', 'N2', 'O2'], 'max': ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3', 'K3', 'L3', 'M3', 'N3', 'O3'], 'sample': ['B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'O4', 'B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5', 'M5', 'N5', 'O5', 'B6', 'C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'M6', 'N6', 'O6', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7', 'J7', 'K7', 'L7', 'M7', 'N7', 'O7', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8', 'N8', 'O8', 'B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9', 'J9', 'K9', 'L9', 'M9', 'N9', 'O9', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10', 'J10', 'K10', 'L10', 'M10', 'N10', 'O10', 'B11', 'C11', 'D11', 'E11', 'F11', 'G11', 'H11', 'I11', 'J11', 'K11', 'L11', 'M11', 'N11', 'O11', 'B12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12', 'I12', 'J12', 'K12', 'L12', 'M12', 'N12', 'O12', 'B13', 'C13', 'D13', 'E13', 'F13', 'G13', 'H13', 'I13', 'J13', 'K13', 'L13', 'M13', 'N13', 'O13', 'B14', 'C14', 'D14', 'E14', 'F14', 'G14', 'H14', 'I14', 'J14', 'K14', 'L14', 'M14', 'N14', 'O14', 'B15', 'C15', 'D15', 'E15', 'F15', 'G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'B16', 'C16', 'D16', 'E16', 'F16', 'G16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16', 'B17', 'C17', 'D17', 'E17', 'F17', 'G17', 'H17', 'I17', 'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'B18', 'C18', 'D18', 'E18', 'F18', 'G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 'B19', 'C19', 'D19', 'E19', 'F19', 'G19', 'H19', 'I19', 'J19', 'K19', 'L19', 'M19', 'N19', 'O19', 'B20', 'C20', 'D20', 'E20', 'F20', 'G20', 'H20', 'I20', 'J20', 'K20', 'L20', 'M20', 'N20', 'O20', 'B21', 'C21', 'D21', 'E21', 'F21', 'G21', 'H21', 'I21', 'J21', 'K21', 'L21', 'M21', 'N21', 'O21', 'B22', 'C22', 'D22', 'E22', 'F22', 'G22', 'H22', 'I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22']}, 'normalised': {'wells': {'A1': -0.08697142857142857, 'B1': -0.024071428571428577, 'C1': -0.07907142857142857, 'D1': -0.08337142857142857, 'E1': -0.04667142857142857, 'F1': -0.011571428571428566, 'G1': 0.017228571428571426, 'H1': 0.001328571428571429, 'I1': 0.03442857142857142, 'J1': 0.014228571428571424, 'K1': -0.06297142857142857, 'L1': -0.06237142857142858, 'M1': 0.14482857142857145, 'N1': 0.2162285714285714, 'O1': 0.08922857142857143, 'P1': 0.31292857142857144, 'A2': -0.08597142857142857, 'P2': 0.10972857142857143, 'A3': -0.08277142857142858, 'P3': 0.007828571428571435, 'A4': -0.08697142857142857, 'P4': -0.06157142857142857, 'A5': -0.08657142857142858, 'P5': -0.04407142857142857, 'A6': -0.08667142857142857, 'P6': 0.019528571428571423, 'A7': -0.08327142857142858, 'P7': 0.02592857142857144, 'A8': -0.08687142857142857, 'P8': -0.004671428571428576, 'A9': -0.08657142857142858, 'P9': -0.018571428571428572, 'A10': -0.08517142857142856, 'P10': 0.02762857142857142, 'A11': -0.08567142857142857, 'P11': 0.05262857142857144, 'A12': -0.08487142857142857, 'P12': 0.03902857142857144, 'A13': -0.08787142857142857, 'P13': -0.019071428571428573, 'A14': -0.08447142857142857, 'P14': 0.035628571428571426, 'A15': -0.08807142857142858, 'P15': -0.005271428571428566, 'A16': -0.08637142857142857, 'P16': 0.06752857142857144, 'A17': -0.07957142857142857, 'P17': 0.7032285714285714, 'A18': -0.08497142857142857, 'P18': 0.8699285714285714, 'A19': -0.08637142857142857, 'P19': 0.8623285714285714, 'A20': -0.07247142857142858, 'P20': 1.0381285714285715, 'A21': -0.05667142857142857, 'P21': 0.1668285714285714, 'A22': -0.028971428571428565, 'P22': 0.06892857142857142, 'A23': -0.07197142857142858, 'B23': -0.04877142857142858, 'C23': -0.06507142857142857, 'D23': -0.06937142857142857, 'E23': -0.018671428571428575, 'F23': 0.013628571428571434, 'G23': 0.05502857142857143, 'H23': 0.010928571428571426, 'I23': 0.019828571428571418, 'J23': 0.04542857142857143, 'K23': -0.02727142857142857, 'L23': -0.051671428571428576, 'M23': 0.07752857142857142, 'N23': 0.05162857142857144, 'O23': 0.8210285714285714, 'P23': 0.9818285714285715, 'A24': -0.08417142857142856, 'B24': -0.04717142857142857, 'C24': -0.05927142857142857, 'D24': -0.06267142857142857, 'E24': 0.05522857142857143, 'F24': -0.008971428571428575, 'G24': 0.00232857142857143, 'H24': 0.05442857142857144, 'I24': 0.03212857142857142, 'J24': 0.1325285714285714, 'K24': -0.06217142857142857, 'L24': -0.06877142857142857, 'M24': 0.16972857142857142, 'N24': 0.055628571428571416, 'O24': 0.6386285714285714, 'P24': 0.7840285714285714, 'B2': -0.05737142857142857, 'C2': -0.06657142857142857, 'D2': -0.07017142857142857, 'E2': -0.051271428571428565, 'F2': -0.05717142857142857, 'G2': -0.05537142857142857, 'H2': -0.02187142857142857, 'I2': -0.021271428571428566, 'J2': 0.017628571428571438, 'K2': -0.03927142857142857, 'L2': -0.05767142857142857, 'M2': 0.1850285714285714, 'N2': 0.24272857142857143, 'O2': 0.05262857142857144, 'B3': 0.14542857142857143, 'C3': -0.04117142857142857, 'D3': -0.04857142857142857, 'E3': 0.06242857142857142, 'F3': 0.19852857142857142, 'G3': 0.018328571428571416, 'H3': 0.19232857142857143, 'I3': -0.002471428571428569, 'J3': 0.23912857142857144, 'K3': -0.04687142857142858, 'L3': -0.05377142857142857, 'M3': -0.007971428571428574, 'N3': 0.09012857142857142, 'O3': -0.052271428571428566, 'B4': 0.25752857142857144, 'C4': 0.09172857142857144, 'D4': 0.08362857142857144, 'E4': 0.04342857142857143, 'F4': 0.27862857142857145, 'G4': 0.03612857142857143, 'H4': 0.14912857142857142, 'I4': 0.11702857142857145, 'J4': 0.16692857142857145, 'K4': -0.03427142857142858, 'L4': 0.019028571428571422, 'M4': -0.019171428571428575, 'N4': 0.12792857142857142, 'O4': 0.021528571428571425, 'B5': 0.2619285714285714, 'C5': 0.8829285714285715, 'D5': 1.3310285714285714, 'E5': 0.061428571428571416, 'F5': 0.25432857142857146, 'G5': -0.0037714285714285645, 'H5': 0.3679285714285715, 'I5': 0.06132857142857143, 'J5': 0.16092857142857145, 'K5': -0.06827142857142857, 'L5': -0.046271428571428574, 'M5': 0.3468285714285714, 'N5': 0.26132857142857147, 'O5': 0.004528571428571437, 'B6': -0.009871428571428559, 'C6': 1.1757285714285715, 'D6': 1.3211285714285714, 'E6': 0.11312857142857144, 'F6': 0.28172857142857144, 'G6': 0.07862857142857144, 'H6': 0.26872857142857143, 'I6': 0.13932857142857144, 'J6': 0.18322857142857144, 'K6': -0.039171428571428565, 'L6': -0.04077142857142857, 'M6': 0.3921285714285715, 'N6': 0.5403285714285715, 'O6': 0.11762857142857144, 'B7': -0.07657142857142857, 'C7': -0.036571428571428574, 'D7': 0.03122857142857144, 'E7': 0.12932857142857143, 'F7': 0.3304285714285714, 'G7': 0.8936285714285714, 'H7': 0.9466285714285714, 'I7': 0.11802857142857145, 'J7': 0.23072857142857142, 'K7': -0.017171428571428574, 'L7': 0.23012857142857143, 'M7': 0.4008285714285714, 'N7': 0.4029285714285714, 'O7': 0.02712857142857142, 'B8': -0.07857142857142857, 'C8': -0.03247142857142857, 'D8': -0.032371428571428565, 'E8': 0.15192857142857144, 'F8': 0.33092857142857146, 'G8': 1.0653285714285714, 'H8': 1.0431285714285714, 'I8': 0.1855285714285714, 'J8': 0.2637285714285714, 'K8': -0.03167142857142857, 'L8': -0.04707142857142857, 'M8': 0.5652285714285714, 'N8': 0.30902857142857143, 'O8': 0.02752857142857143, 'B9': -0.08187142857142857, 'C9': -0.04977142857142858, 'D9': -0.045371428571428576, 'E9': 1.0036285714285715, 'F9': 0.9905285714285714, 'G9': 0.12622857142857144, 'H9': 0.2510285714285714, 'I9': 0.18862857142857145, 'J9': 0.2464285714285714, 'K9': 0.4888285714285715, 'L9': 0.3063285714285714, 'M9': 0.010128571428571431, 'N9': 0.19142857142857142, 'O9': 0.059228571428571436, 'B10': -0.08387142857142857, 'C10': -0.03847142857142857, 'D10': -0.04737142857142858, 'E10': 0.7799285714285714, 'F10': 1.1082285714285716, 'G10': 0.23002857142857144, 'H10': 0.31692857142857145, 'I10': 0.30472857142857146, 'J10': 0.3003285714285714, 'K10': 1.0740285714285716, 'L10': 0.5123285714285715, 'M10': 0.16372857142857142, 'N10': 0.2662285714285715, 'O10': 0.05692857142857144, 'B11': -0.08827142857142857, 'C11': -0.03427142857142858, 'D11': -0.04117142857142857, 'E11': 0.2922285714285714, 'F11': 0.2536285714285714, 'G11': 0.2132285714285714, 'H11': 0.25742857142857145, 'I11': 0.23812857142857144, 'J11': 0.30382857142857145, 'K11': -0.003171428571428575, 'L11': -0.013071428571428567, 'M11': 0.7433285714285714, 'N11': 0.5353285714285715, 'O11': 0.16522857142857142, 'B12': -0.07797142857142858, 'C12': -0.039171428571428565, 'D12': -0.032671428571428573, 'E12': 0.2537285714285714, 'F12': 0.3024285714285714, 'G12': 0.2701285714285714, 'H12': 0.3707285714285714, 'I12': 0.3255285714285714, 'J12': 0.39132857142857147, 'K12': -0.012871428571428575, 'L12': -0.010171428571428581, 'M12': 0.7333285714285714, 'N12': 0.5945285714285714, 'O12': 0.2172285714285714, 'B13': -0.08797142857142856, 'C13': -0.05437142857142857, 'D13': -0.034171428571428575, 'E13': 0.2617285714285714, 'F13': 0.30392857142857144, 'G13': 0.16772857142857142, 'H13': 0.2700285714285714, 'I13': 0.24992857142857142, 'J13': 0.36212857142857147, 'K13': 0.7072285714285714, 'L13': 1.4481285714285714, 'M13': 0.17172857142857142, 'N13': 0.1754285714285714, 'O13': 0.1713285714285714, 'B14': -0.08817142857142857, 'C14': 0.04282857142857144, 'D14': -0.03847142857142857, 'E14': 0.3242285714285714, 'F14': 0.3154285714285714, 'G14': 0.3143285714285714, 'H14': 0.3333285714285714, 'I14': 0.38852857142857145, 'J14': 0.4182285714285715, 'K14': 1.1728285714285716, 'L14': 1.8531285714285715, 'M14': 0.2770285714285714, 'N14': 0.26532857142857147, 'O14': 0.21282857142857145, 'B15': -0.08347142857142857, 'C15': -0.05667142857142857, 'D15': -0.04077142857142857, 'E15': 0.3314285714285714, 'F15': 0.3334285714285714, 'G15': 0.2609285714285714, 'H15': 0.32782857142857147, 'I15': 0.35612857142857146, 'J15': 0.5843285714285714, 'K15': -0.030971428571428566, 'L15': -0.02437142857142857, 'M15': 1.0087285714285714, 'N15': 1.0208285714285714, 'O15': 0.17632857142857142, 'B16': -0.08177142857142858, 'C16': -0.05387142857142857, 'D16': -0.050671428571428576, 'E16': 0.2590285714285714, 'F16': 0.2667285714285714, 'G16': 0.2506285714285714, 'H16': 0.28042857142857147, 'I16': 0.3247285714285715, 'J16': 0.3446285714285714, 'K16': -0.014871428571428577, 'L16': 0.06432857142857143, 'M16': 1.0123285714285715, 'N16': 1.0057285714285715, 'O16': 0.4697285714285715, 'B17': -0.061271428571428574, 'C17': -0.05237142857142857, 'D17': -0.04277142857142857, 'E17': 0.3328285714285715, 'F17': 0.27352857142857145, 'G17': 0.20742857142857143, 'H17': 0.2791285714285714, 'I17': 0.3223285714285714, 'J17': 0.3698285714285715, 'K17': -0.024471428571428575, 'L17': 0.09592857142857142, 'M17': 0.5875285714285715, 'N17': 0.7312285714285714, 'O17': 0.6378285714285714, 'B18': -0.08657142857142858, 'C18': -0.026571428571428565, 'D18': -0.04037142857142857, 'E18': 0.2712285714285714, 'F18': 0.21582857142857145, 'G18': 0.18082857142857142, 'H18': 0.3055285714285715, 'I18': 0.3475285714285714, 'J18': 0.4455285714285715, 'K18': -0.007971428571428574, 'L18': -0.016971428571428568, 'M18': 0.8336285714285715, 'N18': 1.2596285714285715, 'O18': 0.9736285714285715, 'B19': -0.07417142857142857, 'C19': -0.036271428571428566, 'D19': -0.02377142857142857, 'E19': 0.22602857142857144, 'F19': 0.20482857142857144, 'G19': 0.16712857142857143, 'H19': 0.32992857142857146, 'I19': 0.47202857142857146, 'J19': 0.4311285714285714, 'K19': 0.030328571428571427, 'L19': 0.18002857142857145, 'M19': 1.1259285714285714, 'N19': 0.8586285714285714, 'O19': 0.7950285714285714, 'B20': -0.08837142857142857, 'C20': -0.06467142857142857, 'D20': -0.024071428571428577, 'E20': 0.2354285714285714, 'F20': 0.1865285714285714, 'G20': 0.1466285714285714, 'H20': 0.2578285714285714, 'I20': 0.38562857142857143, 'J20': 0.38642857142857145, 'K20': -0.029271428571428573, 'L20': 0.028328571428571425, 'M20': 1.2170285714285713, 'N20': 1.1862285714285714, 'O20': 0.7746285714285714, 'B21': -0.08477142857142858, 'C21': -0.06647142857142857, 'D21': -0.07167142857142857, 'E21': 0.1718285714285714, 'F21': 0.2197285714285714, 'G21': 0.15532857142857145, 'H21': 0.14902857142857143, 'I21': 0.30562857142857147, 'J21': 0.4090285714285714, 'K21': -0.00977142857142857, 'L21': 0.06772857142857142, 'M21': 0.36712857142857147, 'N21': 0.47212857142857145, 'O21': 0.38842857142857146, 'B22': -0.05187142857142857, 'C22': -0.07637142857142856, 'D22': -0.06487142857142857, 'E22': 0.1587285714285714, 'F22': 0.1588285714285714, 'G22': 0.1829285714285714, 'H22': 0.2901285714285714, 'I22': 0.3578285714285714, 'J22': 0.3862285714285715, 'K22': -0.017271428571428576, 'L22': -0.023671428571428565, 'M22': 0.2881285714285714, 'N22': 0.42382857142857144, 'O22': 0.35632857142857144}, 'empty': ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'A2', 'P2', 'A3', 'P3', 'A4', 'P4', 'A5', 'P5', 'A6', 'P6', 'A7', 'P7', 'A8', 'P8', 'A9', 'P9', 'A10', 'P10', 'A11', 'P11', 'A12', 'P12', 'A13', 'P13', 'A14', 'P14', 'A15', 'P15', 'A16', 'P16', 'A17', 'P17', 'A18', 'P18', 'A19', 'P19', 'A20', 'P20', 'A21', 'P21', 'A22', 'P22', 'A23', 'B23', 'C23', 'D23', 'E23', 'F23', 'G23', 'H23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'A24', 'B24', 'C24', 'D24', 'E24', 'F24', 'G24', 'H24', 'I24', 'J24', 'K24', 'L24', 'M24', 'N24', 'O24', 'P24'], 'minimum': ['B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2', 'L2', 'M2', 'N2', 'O2'], 'max': ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3', 'K3', 'L3', 'M3', 'N3', 'O3'], 'sample': ['B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'O4', 'B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5', 'M5', 'N5', 'O5', 'B6', 'C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'M6', 'N6', 'O6', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7', 'J7', 'K7', 'L7', 'M7', 'N7', 'O7', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8', 'N8', 'O8', 'B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9', 'J9', 'K9', 'L9', 'M9', 'N9', 'O9', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10', 'J10', 'K10', 'L10', 'M10', 'N10', 'O10', 'B11', 'C11', 'D11', 'E11', 'F11', 'G11', 'H11', 'I11', 'J11', 'K11', 'L11', 'M11', 'N11', 'O11', 'B12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12', 'I12', 'J12', 'K12', 'L12', 'M12', 'N12', 'O12', 'B13', 'C13', 'D13', 'E13', 'F13', 'G13', 'H13', 'I13', 'J13', 'K13', 'L13', 'M13', 'N13', 'O13', 'B14', 'C14', 'D14', 'E14', 'F14', 'G14', 'H14', 'I14', 'J14', 'K14', 'L14', 'M14', 'N14', 'O14', 'B15', 'C15', 'D15', 'E15', 'F15', 'G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'B16', 'C16', 'D16', 'E16', 'F16', 'G16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16', 'B17', 'C17', 'D17', 'E17', 'F17', 'G17', 'H17', 'I17', 'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'B18', 'C18', 'D18', 'E18', 'F18', 'G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 'B19', 'C19', 'D19', 'E19', 'F19', 'G19', 'H19', 'I19', 'J19', 'K19', 'L19', 'M19', 'N19', 'O19', 'B20', 'C20', 'D20', 'E20', 'F20', 'G20', 'H20', 'I20', 'J20', 'K20', 'L20', 'M20', 'N20', 'O20', 'B21', 'C21', 'D21', 'E21', 'F21', 'G21', 'H21', 'I21', 'J21', 'K21', 'L21', 'M21', 'N21', 'O21', 'B22', 'C22', 'D22', 'E22', 'F22', 'G22', 'H22', 'I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22']}, 'pora': {'wells': {'A1': -175.64916330063474, 'B1': -48.615118291979236, 'C1': -159.6941719561454, 'D1': -168.37853433352566, 'E1': -94.25851125216388, 'F1': -23.369878822850538, 'G1': 34.79515291402193, 'H1': 2.683208309290249, 'I1': 69.53260242354298, 'J1': 28.736295441431036, 'K1': -127.17830351990769, 'L1': -125.96653202538951, 'M1': 292.4985574148875, 'N1': 436.69936526255043, 'O1': 180.20773225620314, 'P1': 631.99653779573, 'A2': -173.62954414310445, 'P2': 221.60992498557414, 'A3': -167.1667628390075, 'P3': 15.810732833237173, 'A4': -175.64916330063474, 'P4': -124.35083669936526, 'A5': -174.84131563762264, 'P5': -89.00750144258511, 'A6': -175.04327755337565, 'P6': 39.44027697634159, 'A7': -168.1765724177727, 'P7': 52.36583958453551, 'A8': -175.44720138488174, 'P8': -9.434506635891529, 'A9': -174.84131563762264, 'P9': -37.50721292556261, 'A10': -172.0138488170802, 'P10': 55.79919215233697, 'A11': -173.02365839584536, 'P11': 106.28967109059438, 'A12': -171.4079630698211, 'P12': 78.82285054818237, 'A13': -177.46682054241202, 'P13': -38.51702250432776, 'A14': -170.600115406809, 'P14': 71.95614541257935, 'A15': -177.87074437391806, 'P15': -10.646278130409684, 'A16': -174.43739180611658, 'P16': 136.38199653779574, 'A17': -160.70398153491058, 'P17': 1420.2538949798038, 'A18': -171.60992498557414, 'P18': 1756.9244085401037, 'A19': -174.43739180611658, 'P19': 1741.5753029428738, 'A20': -146.3646855164455, 'P20': 2096.6243508366997, 'A21': -114.4547028274668, 'P21': 336.93017888055397, 'A22': -58.51125216387766, 'P22': 139.20946335833813, 'A23': -145.35487593768033, 'B23': -98.4997114829775, 'C23': -131.4195037507213, 'D23': -140.10386612810157, 'E23': -37.709174841315644, 'F23': 27.52452394691288, 'G23': 111.13675706866705, 'H23': 22.07155222158107, 'I23': 40.04616272360067, 'J23': 91.74841315637623, 'K23': -55.07789959607617, 'L23': -104.35660703981536, 'M23': 156.57818811309866, 'N23': 104.27005193306407, 'O23': 1658.1650317368726, 'P23': 1982.919792267744, 'A24': -169.99422965954992, 'B24': -95.26832083092903, 'C24': -119.7057126370456, 'D24': -126.57241777264859, 'E24': 111.54068090017311, 'F24': -18.11886901327179, 'G24': 4.702827466820545, 'H24': 109.9249855741489, 'I24': 64.8874783612233, 'J24': 267.65724177726486, 'K24': -125.56260819388345, 'L24': -138.89209463358338, 'M24': 342.78707443739177, 'N24': 112.3485285631852, 'O24': 1289.7864974033469, 'P24': 1583.4391229082516, 'B2': -115.86843623773804, 'C2': -134.44893248701675, 'D2': -141.7195614541258, 'E2': -103.54875937680322, 'F2': -115.46451240623198, 'G2': -111.82919792267745, 'H2': -44.17195614541257, 'I2': -42.9601846508944, 'J2': 35.60300057703407, 'K2': -79.3133294864397, 'L2': -116.47432198499712, 'M2': 373.6872475476053, 'N2': 490.21927293710337, 'O2': 106.28967109059438, 'B3': 293.7103289094057, 'C3': -83.15060588574725, 'D3': -98.09578765147144, 'E3': 126.08193883439122, 'F3': 400.9521061742643, 'G3': 37.01673398730523, 'H3': 388.4304673975765, 'I3': -4.991344489324865, 'J3': 482.9486439699943, 'K3': -94.66243508366995, 'L3': -108.59780727062896, 'M3': -16.099249855741494, 'N3': 182.02538949798037, 'O3': -105.56837853433352, 'B4': 520.1096364685517, 'C4': 185.25678015002887, 'D4': 168.89786497403352, 'E4': 87.70917484131563, 'F4': 562.7236006924409, 'G4': 72.96595499134449, 'H4': 301.18291979226774, 'I4': 236.35314483554535, 'J4': 337.13214079630706, 'K4': -69.21523369878825, 'L4': 38.43046739757644, 'M4': -38.718984420080794, 'N4': 258.3669936526255, 'O4': 43.47951529140219, 'B5': 528.9959607616848, 'C5': 1783.1794575879978, 'D5': 2688.1708020773226, 'E5': 124.06231967686091, 'F5': 513.6468551644548, 'G5': -7.616849394114239, 'H5': 743.0755914598964, 'I5': 123.8603577611079, 'J5': 325.0144258511253, 'K5': -137.88228505481823, 'L5': -93.45066358915176, 'M5': 700.4616272360068, 'N5': 527.7841892671669, 'O5': 9.145989613387208, 'B6': -19.936526255049024, 'C6': 2374.523946912868, 'D6': 2668.1765724177726, 'E6': 228.47663012117718, 'F6': 568.9844200807848, 'G6': 158.79976918638204, 'H6': 542.729371032891, 'I6': 281.3906520484709, 'J6': 370.0519330640508, 'K6': -79.11136757068665, 'L6': -82.34275822273513, 'M6': 791.9503750721293, 'N6': 1091.257934218119, 'O6': 237.56491633006354, 'B7': -154.64512406231967, 'C7': -73.86035776110792, 'D7': 63.06982111944607, 'E7': 261.19446047316796, 'F7': 667.33987305251, 'G7': 1804.7893825735719, 'H7': 1911.8291979226774, 'I7': 238.37276399307567, 'J7': 465.98384304673976, 'K7': -34.6797461050202, 'L7': 464.77207155222163, 'M7': 809.5210617426428, 'N7': 813.7622619734565, 'O7': 54.78938257357182, 'B8': -158.68436237738027, 'C8': -65.5799192152337, 'D8': -65.37795729948066, 'E8': 306.8378534333526, 'F8': 668.3496826312754, 'G8': 2151.5579919215234, 'H8': 2106.722446624351, 'I8': 374.6970571263704, 'J8': 532.6312752452395, 'K8': -63.96422388920947, 'L8': -95.066358915176, 'M8': 1141.5464512406231, 'N8': 624.1200230813619, 'O8': 55.59723023658397, 'B9': -165.34910559723022, 'C9': -100.5193306405078, 'D9': -91.6330063473745, 'E9': 2026.9474899019044, 'F9': 2000.4904789382572, 'G9': 254.93364108482405, 'H9': 506.9821119446047, 'I9': 380.9578765147144, 'J9': 497.69186381996536, 'K9': 987.2475476053089, 'L9': 618.66705135603, 'M9': 20.455856895556845, 'N9': 386.6128101557992, 'O9': 119.6191575302943, 'B10': -169.38834391229082, 'C10': -77.69763416041548, 'D10': -95.67224466243509, 'E10': 1575.1586843623775, 'F10': 2238.1996537795735, 'G10': 464.5701096364686, 'H10': 640.0750144258511, 'I10': 615.4356607039816, 'J10': 606.5493364108482, 'K10': 2169.1286785920374, 'L10': 1034.7085978072707, 'M10': 330.66935949221005, 'N10': 537.6803231390654, 'O10': 114.97403346797464, 'B11': -178.27466820542412, 'C11': -69.21523369878825, 'D11': -83.15060588574725, 'E11': 590.1904212348528, 'F11': 512.2331217541835, 'G11': 430.6405077899596, 'H11': 519.9076745527987, 'I11': 480.929024812464, 'J11': 613.6180034622043, 'K11': -6.405077899596083, 'L11': -26.39930755914598, 'M11': 1501.2406231967689, 'N11': 1081.1598384304675, 'O11': 333.69878822850546, 'B12': -157.4725908828621, 'C12': -79.11136757068665, 'D12': -65.98384304673976, 'E12': 512.4350836699366, 'F12': 610.7905366416618, 'G12': 545.5568378534333, 'H12': 748.7305251009809, 'I12': 657.4437391806115, 'J12': 790.3346797461052, 'K12': -25.99538372763994, 'L12': -20.542412002308158, 'M12': 1481.044431621466, 'N12': 1200.7212925562608, 'O12': 438.71898442008074, 'B13': -177.668782458165, 'C13': -109.80957876514714, 'D13': -69.01327178303521, 'E13': 528.5920369301789, 'F13': 613.8199653779573, 'G13': 338.7478361223312, 'H13': 545.3548759376803, 'I13': 504.76053087132135, 'J13': 731.3618003462205, 'K13': 1428.3323716099248, 'L13': 2924.6682054241205, 'M13': 346.8263127524524, 'N13': 354.29890363531445, 'O13': 346.01846508944027, 'B14': -178.0727062896711, 'C14': 86.49740334679748, 'D14': -77.69763416041548, 'E14': 654.8182342758222, 'F14': 637.0455856895557, 'G14': 634.8240046162723, 'H14': 673.1967686093479, 'I14': 784.6797461050203, 'J14': 844.6624350836702, 'K14': 2368.6670513560302, 'L14': 3742.6139642238895, 'M14': 559.4922100403924, 'N14': 535.862665897288, 'O14': 429.8326601269476, 'B15': -168.58049624927872, 'C15': -114.4547028274668, 'D15': -82.34275822273513, 'E15': 669.3594922100403, 'F15': 673.398730525101, 'G15': 526.9763416041546, 'H15': 662.0888632429314, 'I15': 719.2440854010388, 'J15': 1180.121177149452, 'K15': -62.55049047893825, 'L15': -49.22100403923832, 'M15': 2037.247547605309, 'N15': 2061.684939411425, 'O15': 356.11656087709173, 'B16': -165.14714368147722, 'C16': -108.799769186382, 'D16': -102.33698788228507, 'E16': 523.139065204847, 'F16': 538.6901327178304, 'G16': 506.17426428159257, 'H16': 566.3589151759955, 'I16': 655.8280438545876, 'J16': 696.0184650894403, 'K16': -30.034622042700533, 'L16': 129.9192152336988, 'M16': 2044.5181765724178, 'N16': 2031.188690132718, 'O16': 948.6728216964804, 'B17': -123.74495095210618, 'C17': -105.77034045008655, 'D17': -86.38199653779574, 'E17': 672.1869590305829, 'F17': 552.4235429890365, 'G17': 418.92671667628395, 'H17': 563.7334102712059, 'I17': 650.9809578765147, 'J17': 746.9128678592039, 'K17': -49.422965954991355, 'L17': 193.73918061165608, 'M17': 1186.583958453549, 'N17': 1476.8032313906522, 'O17': 1288.1708020773226, 'B18': -174.84131563762264, 'C18': -53.66416618580495, 'D18': -81.53491055972303, 'E18': 547.7784189267165, 'F18': 435.8915175995384, 'G18': 365.2048470859781, 'H18': 617.051356030006, 'I18': 701.8753606462781, 'J18': 899.7980380842472, 'K18': -16.099249855741494, 'L18': -34.27582227351413, 'M18': 1683.6122331217543, 'N18': 2543.96999422966, 'O18': 1966.3589151759957, 'B19': -149.79803808424697, 'C19': -73.2544720138488, 'D19': -48.009232544720135, 'E19': 456.4916330063474, 'F19': 413.67570686670524, 'G19': 337.53606462781306, 'H19': 666.3300634737451, 'I19': 953.3179457588, 'J19': 870.7155222158108, 'K19': 61.252163877668785, 'L19': 363.5891517599539, 'M19': 2273.9469128678593, 'N19': 1734.1027120600115, 'O19': 1605.6549336410849, 'B20': -178.47663012117718, 'C20': -130.61165608770918, 'D20': -48.615118291979236, 'E20': 475.47605308713213, 'F20': 376.7166762839007, 'G20': 296.133871898442, 'H20': 520.7155222158107, 'I20': 778.8228505481824, 'J20': 780.4385458742066, 'K20': -59.117137911136766, 'L20': 57.21292556260819, 'M20': 2457.934218118869, 'N20': 2395.7299480669362, 'O20': 1564.454702827467, 'B21': -171.20600115406813, 'C21': -134.24697057126372, 'D21': -144.74899019042124, 'E21': 347.0282746682054, 'F21': 443.7680323139065, 'G21': 313.70455856895563, 'H21': 300.9809578765147, 'I21': 617.2533179457589, 'J21': 826.0819388343913, 'K21': -19.734564339296018, 'L21': 136.78592036930178, 'M21': 741.459896133872, 'N21': 953.5199076745529, 'O21': 784.4777841892674, 'B22': -104.76053087132141, 'C22': -154.2412002308136, 'D22': -131.01557991921524, 'E22': 320.57126370455853, 'F22': 320.77322562031156, 'G22': 369.4460473167917, 'H22': 585.9492210040393, 'I22': 722.6774379688401, 'J22': 780.0346220427007, 'K22': -34.88170802077324, 'L22': -47.807270628967096, 'M22': 581.9099826889786, 'N22': 855.9723023658397, 'O22': 719.6480092325447}, 'empty': ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'A2', 'P2', 'A3', 'P3', 'A4', 'P4', 'A5', 'P5', 'A6', 'P6', 'A7', 'P7', 'A8', 'P8', 'A9', 'P9', 'A10', 'P10', 'A11', 'P11', 'A12', 'P12', 'A13', 'P13', 'A14', 'P14', 'A15', 'P15', 'A16', 'P16', 'A17', 'P17', 'A18', 'P18', 'A19', 'P19', 'A20', 'P20', 'A21', 'P21', 'A22', 'P22', 'A23', 'B23', 'C23', 'D23', 'E23', 'F23', 'G23', 'H23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'A24', 'B24', 'C24', 'D24', 'E24', 'F24', 'G24', 'H24', 'I24', 'J24', 'K24', 'L24', 'M24', 'N24', 'O24', 'P24'], 'minimum': ['B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2', 'L2', 'M2', 'N2', 'O2'], 'max': ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3', 'K3', 'L3', 'M3', 'N3', 'O3'], 'sample': ['B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'O4', 'B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5', 'M5', 'N5', 'O5', 'B6', 'C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'M6', 'N6', 'O6', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7', 'J7', 'K7', 'L7', 'M7', 'N7', 'O7', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8', 'N8', 'O8', 'B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9', 'J9', 'K9', 'L9', 'M9', 'N9', 'O9', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10', 'J10', 'K10', 'L10', 'M10', 'N10', 'O10', 'B11', 'C11', 'D11', 'E11', 'F11', 'G11', 'H11', 'I11', 'J11', 'K11', 'L11', 'M11', 'N11', 'O11', 'B12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12', 'I12', 'J12', 'K12', 'L12', 'M12', 'N12', 'O12', 'B13', 'C13', 'D13', 'E13', 'F13', 'G13', 'H13', 'I13', 'J13', 'K13', 'L13', 'M13', 'N13', 'O13', 'B14', 'C14', 'D14', 'E14', 'F14', 'G14', 'H14', 'I14', 'J14', 'K14', 'L14', 'M14', 'N14', 'O14', 'B15', 'C15', 'D15', 'E15', 'F15', 'G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'B16', 'C16', 'D16', 'E16', 'F16', 'G16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16', 'B17', 'C17', 'D17', 'E17', 'F17', 'G17', 'H17', 'I17', 'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'B18', 'C18', 'D18', 'E18', 'F18', 'G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 'B19', 'C19', 'D19', 'E19', 'F19', 'G19', 'H19', 'I19', 'J19', 'K19', 'L19', 'M19', 'N19', 'O19', 'B20', 'C20', 'D20', 'E20', 'F20', 'G20', 'H20', 'I20', 'J20', 'K20', 'L20', 'M20', 'N20', 'O20', 'B21', 'C21', 'D21', 'E21', 'F21', 'G21', 'H21', 'I21', 'J21', 'K21', 'L21', 'M21', 'N21', 'O21', 'B22', 'C22', 'D22', 'E22', 'F22', 'G22', 'H22', 'I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22']}}, 'calculations': {'original': {'empty': {'avg': 0.20203, 'stdev': 0.25698448143152386, 'pstdev': 0.25555280187163754, 'pvariance': 0.06530723454444444, 'variance': 0.06604102369662922, 'st_dev_%': 127.20114905287525}, 'minimum': {'avg': 0.13677142857142857, 'stdev': 0.09752432776233069, 'pstdev': 0.09397679370515914, 'pvariance': 0.00883163775510204, 'variance': 0.009510994505494506, 'st_dev_%': 71.30460563362386}, 'max': {'avg': 0.18628571428571428, 'stdev': 0.10550736654662742, 'pstdev': 0.10166944236201989, 'pvariance': 0.010336675510204082, 'variance': 0.011131804395604395, 'st_dev_%': 56.63739001736135}, 'sample': {'avg': 0.40891315789473687, 'stdev': 0.3422797545203607, 'pstdev': 0.3416357657005735, 'pvariance': 0.11671499640581717, 'variance': 0.11715543035451838, 'st_dev_%': 83.70475439884743}, 'other': {'S/B': 1.3620221433047837}}, 'normalised': {'empty': {'avg': 0.06525857142857143, 'stdev': 0.25698448143152386, 'pstdev': 0.2555528018716376, 'pvariance': 0.06530723454444445, 'variance': 0.06604102369662922, 'st_dev_%': 393.79421860744446}, 'minimum': {'avg': 1.982541115402065e-18, 'stdev': 0.09752432776233069, 'pstdev': 0.09397679370515914, 'pvariance': 0.00883163775510204, 'variance': 0.009510994505494506, 'st_dev_%': 4.919157893103896e+18}, 'max': {'avg': 0.04951428571428571, 'stdev': 0.10550736654662742, 'pstdev': 0.10166944236201989, 'pvariance': 0.010336675510204082, 'variance': 0.011131804395604395, 'st_dev_%': 213.08469873813965}, 'sample': {'avg': 0.27214172932330827, 'stdev': 0.3422797545203607, 'pstdev': 0.3416357657005735, 'pvariance': 0.11671499640581719, 'variance': 0.11715543035451838, 'st_dev_%': 125.77260950441284}, 'other': {'S/B': 2.4975162093545824e+16}}, 'pora': {'empty': {'avg': 131.79746105020197, 'stdev': 519.0107818870938, 'pstdev': 516.1193344205029, 'pvariance': 266379.1673626629, 'variance': 269372.1917150524, 'st_dev_%': 393.7942186074444}, 'minimum': {'avg': 7.105427357601002e-15, 'stdev': 196.96200067406662, 'pstdev': 189.7973329302118, 'pvariance': 36023.02758742166, 'variance': 38794.02970953102, 'st_dev_%': 2.7719937276308557e+18}, 'max': {'avg': 100.00000000000001, 'stdev': 213.08469873813965, 'pstdev': 205.3335535297574, 'pvariance': 42161.86820515775, 'variance': 45405.08883632373, 'st_dev_%': 213.0846987381396}, 'sample': {'avg': 549.6226501047772, 'stdev': 691.2747494640869, 'pstdev': 689.9741373064094, 'pvariance': 476064.31015172385, 'variance': 477860.779246636, 'st_dev_%': 125.77260950441287}, 'other': {'S/B': 1.4073748835532802e+16}}, 'other': {'z_prime': -11.30140098236619}}}}

    analyse_method = None
    output_file = "C:/Users/phch/Desktop/more_data_files/excel_test_output/test.xlsx"
    bio_final_report_controller(analyse_method, all_plate_data, output_file, final_report_setup)